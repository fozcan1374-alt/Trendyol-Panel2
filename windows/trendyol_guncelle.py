#!/usr/bin/env python3
"""TRENDYOL GÜNCELLE — Downloads'taki EN YENİ tarife dosyalarını (normal + Plus, içerikten
tanır) Trendyol_Karlilik.xlsx'e işler, API fiyat/stok senkronunu çalıştırır, Excel'i açar.
Kullanım: dosyaları Downloads'a indir → bunu çalıştır (veya 'Trendyol Güncelle.command' çift tık)."""
import openpyxl, glob, os, warnings, subprocess, sys
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import ayar          # yollar paketin konumundan türetilir
warnings.filterwarnings('ignore')

ANA=ayar.XLSX
NAVY='1F4E78'

def en_yeni_dosyalar():
    normal=plus=None; nt=pt=0
    for f in glob.glob(os.path.join(ayar.INDIRILENLER, ayar.satici_no()+'-*.xlsx')):
        try:
            wb=openpyxl.load_workbook(f, read_only=True)
            sh=wb.sheetnames[0]; m=os.path.getmtime(f)
            if sh=='TyPlusÜrünleri' and m>pt: plus, pt=f, m
            elif sh=='KomisyonTarifeleriÜrünleri' and m>nt: normal, nt=f, m
            wb.close()
        except Exception: pass
    return normal, plus

def sekmeye_yaz(wb, sekme_adi, kaynak_yol, beklenen_baslik, beklenen_kolon):
    src=openpyxl.load_workbook(kaynak_yol, data_only=True)
    ws_s=src[src.sheetnames[0]]
    if ws_s.max_column!=beklenen_kolon or ws_s.cell(1,1).value!=beklenen_baslik:
        print(f"⚠ {os.path.basename(kaynak_yol)}: format beklenenden farklı "
              f"({ws_s.max_column} kolon, A1='{ws_s.cell(1,1).value}') — İŞLENMEDİ. Format değiştiyse söyle.")
        return False
    ws=wb[sekme_adi]
    for row in ws.iter_rows():
        for c in row: c.value=None
    for r in range(1, ws_s.max_row+1):
        for c in range(1, ws_s.max_column+1):
            ws.cell(r,c, ws_s.cell(r,c).value)
    for c in range(1, ws_s.max_column+1):
        cell=ws.cell(1,c); cell.font=Font(bold=True,color='FFFFFF',size=10)
        cell.fill=PatternFill('solid',fgColor=NAVY); cell.alignment=Alignment(horizontal='center',wrap_text=True)
    print(f"✓ {sekme_adi} ← {os.path.basename(kaynak_yol)} ({ws_s.max_row-1} ürün)")
    return True

normal, plus = en_yeni_dosyalar()
wb=openpyxl.load_workbook(ANA)
if 'Plus Tarifeleri' not in wb.sheetnames:
    wb.create_sheet('Plus Tarifeleri')
if normal: sekmeye_yaz(wb,'Komisyon Tarifeleri',normal,'ÜRÜN İSMİ',35)
else: print('⚠ Downloads\'ta normal tarife dosyası (KomisyonTarifeleriÜrünleri) bulunamadı')
if plus: sekmeye_yaz(wb,'Plus Tarifeleri',plus,'Ürün İsmi',31)
else: print('⚠ Downloads\'ta Plus tarife dosyası (TyPlusÜrünleri) bulunamadı')

# Maliyetler'de olmayan yeni barkodları ekle
kt=wb['Komisyon Tarifeleri']; ml=wb['Maliyetler']
mevcut={str(ml.cell(r,1).value).strip() for r in range(2, ml.max_row+1) if ml.cell(r,1).value}
sari=PatternFill('solid',fgColor='FFF2CC'); yeni=0
for rr in range(2, kt.max_row+1):
    bk=kt.cell(rr,2).value
    if bk is None: continue
    bs=str(bk).strip()
    if bs in mevcut: continue
    r=ml.max_row+1
    ml.cell(r,1,bs); ml.cell(r,2,kt.cell(rr,1).value)
    for cc in (3,4,5): ml.cell(r,cc).fill=sari
    mevcut.add(bs); yeni+=1
if yeni: print(f"✓ Maliyetler'e {yeni} yeni ürün eklendi (maliyetlerini gir — sarı)")
wb.save(ANA)
print('✓ kaydedildi:', ANA)
# API fiyat/stok senkronu
try:
    subprocess.run([sys.executable,ayar.SYNC_PY],check=True)
except Exception as e: print('⚠ API senkron hatası:',e)
if '--no-open' not in sys.argv:
    subprocess.run(['osascript','-e','tell application "Microsoft Excel" to close (every workbook whose name contains "Trendyol_Karlilik") saving no'],capture_output=True)
    subprocess.run(['open',ANA])
