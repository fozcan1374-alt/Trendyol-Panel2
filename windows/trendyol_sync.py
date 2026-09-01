#!/usr/bin/env python3
"""TRENDYOL API SENKRON — Trendyol_Karlilik.xlsx'i canlı fiyat/stokla besler.
Kullanım: python3 ~/Desktop/trendyol_sync.py
API bilgileri: ~/.trendyol_api.json (chmod 600). Sadece OKUMA yapar, panele yazmaz."""
import requests, json, base64, datetime, warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import ayar          # yollar paketin konumundan türetilir
warnings.filterwarnings('ignore')

cfg=json.load(open(ayar.TY_API))
sid=cfg['seller_id']
tok=base64.b64encode(f"{cfg['api_key']}:{cfg['api_secret']}".encode()).decode()
H={'Authorization':f'Basic {tok}','User-Agent':f'{sid} - SelfIntegration','Accept':'application/json'}
BASE=f'https://apigw.trendyol.com/integration/product/sellers/{sid}/products/approved/inventory-and-price'

fiyatlar={}; page=0
while True:
    r=requests.get(BASE,params={'page':page,'size':100},headers=H,timeout=30); r.raise_for_status()
    js=r.json()
    for p in js.get('content',[]):
        for v in p.get('variants',[]):
            bk=str(v.get('barcode','')).strip()
            if bk: fiyatlar[bk]=(v.get('salePrice'), v.get('quantity'))
    if page>=js.get('totalPages',1)-1: break
    page+=1
print(f'API: {len(fiyatlar)} barkod çekildi ({page+1} sayfa)')

yol=ayar.XLSX
wb=openpyxl.load_workbook(yol)
ml=wb['Maliyetler']; hs=wb['KÂRLILIK']
NAVY='1F4E78'; HF=PatternFill('solid',fgColor=NAVY); HFONT=Font(bold=True,color='FFFFFF',size=10)
# Maliyetler F/G/H başlıkları
for c,h,w in [(6,'Canlı Satış Fiyatı (API)',20),(7,'Canlı Stok (API)',14),(8,'API Güncelleme',18)]:
    cell=ml.cell(1,c)
    if cell.value!=h:
        cell.value=h; cell.font=HFONT; cell.fill=HF; cell.alignment=Alignment(horizontal='center',wrap_text=True)
        ml.column_dimensions[openpyxl.utils.get_column_letter(c)].width=w
simdi=datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
n=0
for r in range(2, ml.max_row+1):
    bk=ml.cell(r,1).value
    if not bk: continue
    es=fiyatlar.get(str(bk).strip())
    if es:
        ml.cell(r,6,es[0]); ml.cell(r,7,es[1]); ml.cell(r,8,simdi); n+=1
# KÂRLILIK'a canlı fiyat/stok lookup kolonları (AJ=36, AK=37) — bir kez kurulur
if hs.cell(1,36).value!='CANLI FİYAT (API)':
    for c,h in [(36,'CANLI FİYAT (API)'),(37,'CANLI STOK')]:
        cell=hs.cell(1,c); cell.value=h; cell.font=HFONT; cell.fill=HF; cell.alignment=Alignment(horizontal='center',wrap_text=True)
        hs.column_dimensions[openpyxl.utils.get_column_letter(c)].width=13
    for i in range(2,302):
        hs.cell(i,36,f'=IF($B{i}="","",IFERROR(VLOOKUP($B{i},Maliyetler!$A:$G,6,FALSE),""))').number_format='#,##0.00'
        hs.cell(i,37,f'=IF($B{i}="","",IFERROR(VLOOKUP($B{i},Maliyetler!$A:$G,7,FALSE),""))')
wb.save(yol)
print(f'✓ Excel güncellendi: {n} ürün eşleşti | {simdi}')
# eşleşen ürünlerin özeti
wb2=openpyxl.load_workbook(yol, data_only=True)
ml2=wb2['Maliyetler']
for r in range(2, ml2.max_row+1):
    if ml2.cell(r,6).value is not None:
        print(f"  {str(ml2.cell(r,2).value)[:45]:<46} canlı fiyat: {ml2.cell(r,6).value:>10} | stok: {ml2.cell(r,7).value}")
