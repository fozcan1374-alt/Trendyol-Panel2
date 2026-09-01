#!/usr/bin/env python3
"""FATURA → MALİYET — Downloads'taki e-fatura PDF'lerini okur, kalemleri Trendyol ürünleriyle
eşleştirir (barkod VEYA contentId), maliyeti (KDV dahil birim fiyat) Excel'e yazar.
Eşleşme: fatura kalem kodu = EAN barkod ya da Trendyol contentId.
Kullanım: python3 fatura_maliyet.py [--yontem son|ortalama]"""
import fitz, re, json, base64, glob, os, sys, datetime, warnings
import requests, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import ayar          # yollar paketin konumundan türetilir
warnings.filterwarnings('ignore')

XLSX=ayar.XLSX
KDV=1.20
YONTEM='ortalama' if '--yontem' in sys.argv and 'ortalama' in sys.argv else 'son'

KALEM=re.compile(r'\n\s*\d{1,2}\n\s*([0-9A-Za-z][0-9A-Za-z\-_.]*)\s*/\s*(.*?)\n\s*(\d+(?:[.,]\d+)?)\n\s*([\d.,]+)TL', re.S)

def _sayi(x):
    try: return float(str(x).replace('.','').replace(',','.'))
    except Exception: return None

def pdf_kalemleri(yol):
    out=[]
    try: d=fitz.open(yol)
    except Exception: return out
    for i,p in enumerate(d):
        t=p.get_text()
        if 'e-FATURA' not in t and 'Mal Hizmet' not in t: continue
        fno=re.search(r'Fatura No:\s*\n?\s*(\S+)',t)
        ftar=re.search(r'Fatura Tarihi:\s*\n?\s*(\S+)',t)
        for m in KALEM.finditer(t):
            kod, ad, mik, bf = m.groups()
            mik=_sayi(mik); bf=_sayi(bf)
            if not mik or not bf or mik<=0 or bf<=0: continue
            tar=ftar.group(1) if ftar else ''
            try: ts=datetime.datetime.strptime(tar,'%d-%m-%Y')
            except Exception: ts=datetime.datetime(1970,1,1)
            out.append(dict(kod=kod, ad=' '.join(ad.split())[:60], adet=mik,
                            birim=round(bf*KDV,2), tarih=tar, ts=ts,
                            fatura=fno.group(1) if fno else '', dosya=os.path.basename(yol)))
    return out

def trendyol_haritasi():
    cfg=json.load(open(ayar.TY_API))
    sid=cfg['seller_id']
    tok=base64.b64encode(f"{cfg['api_key']}:{cfg['api_secret']}".encode()).decode()
    H={'Authorization':f'Basic {tok}','User-Agent':f'{sid} - SelfIntegration'}
    bar={}; cid={}
    for page in range(20):
        r=requests.get(f'https://apigw.trendyol.com/integration/product/sellers/{sid}/products/approved/inventory-and-price',
                       params={'page':page,'size':100},headers=H,timeout=30).json()
        for p in r.get('content',[]):
            for v in p.get('variants',[]):
                b=str(v.get('barcode','')).strip()
                if b: bar[b]=b; cid[str(p.get('contentId'))]=b
        if page>=r.get('totalPages',1)-1: break
    return bar, cid

def calistir(klasor=ayar.INDIRILENLER, log=print):
    dosyalar=sorted(glob.glob(os.path.join(klasor,'*.pdf')), key=os.path.getmtime, reverse=True)[:40]
    kalemler=[]
    for f in dosyalar: kalemler+=pdf_kalemleri(f)
    log(f'{len(dosyalar)} PDF tarandı → {len(kalemler)} fatura kalemi')
    if not kalemler: return {'yazilan':0,'eslesen':0,'kalem':0,'satirlar':[]}
    bar,cid=trendyol_haritasi()
    grup={}
    for k in kalemler:
        b=bar.get(k['kod']) or cid.get(k['kod'])
        if not b: continue
        grup.setdefault(b,[]).append(k)
    log(f'eşleşen ürün: {len(grup)}')
    # maliyet hesabı
    maliyet={}
    for b,ks in grup.items():
        ks.sort(key=lambda x:x['ts'])
        if YONTEM=='ortalama':
            top=sum(x['birim']*x['adet'] for x in ks); ad=sum(x['adet'] for x in ks)
            m=round(top/ad,2) if ad else ks[-1]['birim']
        else:
            m=ks[-1]['birim']
        maliyet[b]=dict(maliyet=m, ad=ks[-1]['ad'], tarih=ks[-1]['tarih'],
                        fatura=ks[-1]['fatura'], n=len(ks))
    # Excel'e yaz
    wb=openpyxl.load_workbook(XLSX)
    ml=wb['Maliyetler']
    if ml.cell(1,9).value!='Maliyet Kaynağı':
        c=ml.cell(1,9,'Maliyet Kaynağı'); c.font=Font(bold=True,color='FFFFFF',size=10)
        c.fill=PatternFill('solid',fgColor='1F4E78'); c.alignment=Alignment(horizontal='center',wrap_text=True)
        ml.column_dimensions['I'].width=28
    satir={}
    for r in range(2, ml.max_row+1):
        b=ml.cell(r,1).value
        if b: satir[str(b).strip()]=r
    yaz=0; degisen=[]
    for b,d in maliyet.items():
        r=satir.get(b)
        if r is None:
            r=ml.max_row+1; ml.cell(r,1,b); ml.cell(r,2,d['ad']); satir[b]=r
        eski=ml.cell(r,3).value
        if eski is None or abs(float(eski)-d['maliyet'])>0.01:
            ml.cell(r,3,d['maliyet'])
            degisen.append(dict(barkod=b, ad=d['ad'], eski=eski, yeni=d['maliyet'], tarih=d['tarih']))
            yaz+=1
        ml.cell(r,9,f"fatura {d['fatura']} · {d['tarih']} · {YONTEM}")
    wb.save(XLSX)
    log(f'Excel: {yaz} üründe maliyet güncellendi ({YONTEM} alış, KDV dahil)')
    for d in degisen[:25]:
        log(f"   {d['ad'][:38]:<40} {('—' if d['eski'] is None else format(float(d['eski']),',.2f')):>10} → {d['yeni']:>10,.2f}  ({d['tarih']})")
    return {'yazilan':yaz,'eslesen':len(grup),'kalem':len(kalemler),'satirlar':degisen}

if __name__=='__main__':
    calistir()
