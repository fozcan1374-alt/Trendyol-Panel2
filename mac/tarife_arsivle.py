#!/usr/bin/env python3
"""TARİFE ARŞİVİ — Downloads'a inen Trendyol panel export'larını hafta hafta toplar.

Dosya adı Trendyol tarafında hep `<satici-no>-<tarih>-<saat>.xlsx` olduğu için hangi ekranın
export'u olduğu ADINDAN anlaşılmıyor; SEKME ADINDAN tanınıyor. Yeni bir ekran eklenirse
sadece TIP sözlüğüne bir satır ekle.

    python3 tarife_arsivle.py            → Downloads'takileri arşive kopyala
    python3 tarife_arsivle.py --liste    → arşivin içeriğini yaz

Hedef: ~/Desktop/Trendyol/Tarifeler/<yıl>-H<ISO hafta>/<tip>_<tarih>_<saat>.xlsx
Kopyalar, taşımaz — Downloads'taki orijinal durur.
"""
import openpyxl, glob, os, shutil, datetime, sys, warnings
import ayar          # yollar paketin konumundan türetilir
warnings.filterwarnings('ignore')

KOK = ayar.TARIFE_ARSIV
KAYNAK = ayar.INDIRILENLER

# sekme adı → arşiv tipi
TIP = {
    'KomisyonTarifeleriÜrünleri': 'komisyon',   # barem komisyon tarifesi
    'TyPlusÜrünleri':             'typlus',     # TY Plus teklifleri
    'YıldızlıÜrünEtiketleri':     'yildiz',     # avantajlı/yıldızlı ürün eşikleri
    'TeklifÜrünleri':             'flas',       # flaş ürün (24 saat / 3 saat) slotları
}
ADI = {'komisyon': 'Komisyon Tarifesi', 'typlus': 'TY Plus',
       'yildiz': 'Yıldızlı Ürün', 'flas': 'Flaş Ürün'}


def tani(yol):
    """Dosyayı sekme adından tanır → (tip, sekme). Tanınmazsa (None, ilk sekme)."""
    try:
        wb = openpyxl.load_workbook(yol, read_only=True)
    except Exception:
        return None, None
    for sh in wb.sheetnames:
        if sh in TIP:
            return TIP[sh], sh
    return None, (wb.sheetnames[0] if wb.sheetnames else None)


def arsivle(log=print):
    yeni = []
    for f in sorted(glob.glob(os.path.join(KAYNAK, ayar.satici_no()+'-*.xlsx'))):
        t, sh = tani(f)
        if not t:
            log(f'  ? tanınmadı: {os.path.basename(f)}  (sekme: {sh})')
            continue
        mt = datetime.datetime.fromtimestamp(os.path.getmtime(f))
        hafta = mt.strftime('%Y-H%V')                 # ISO hafta numarası
        hedef = os.path.join(KOK, hafta)
        os.makedirs(hedef, exist_ok=True)
        yol = os.path.join(hedef, f'{t}_{mt:%Y-%m-%d_%H%M}.xlsx')
        if os.path.exists(yol):
            continue
        # aynı dakikada iki kopya inmişse (tarayıcı "(1)" ekliyor) tekrar yazma
        shutil.copy2(f, yol)
        yeni.append((hafta, t, os.path.basename(yol)))
    for h, t, a in yeni:
        log(f'  ✓ {h}/{a}   [{ADI[t]}]')
    log(f'{len(yeni)} yeni dosya arşivlendi' if yeni else 'yeni dosya yok')
    return yeni


def liste(log=print):
    if not os.path.isdir(KOK):
        log('arşiv yok'); return
    for h in sorted(os.listdir(KOK)):
        p = os.path.join(KOK, h)
        if not os.path.isdir(p):
            continue
        d = sorted(os.listdir(p))
        log(f'\n■ {h}  ({len(d)} dosya)')
        for x in d:
            t = x.split('_')[0]
            log(f'    {ADI.get(t, t):<20} {x}')


if __name__ == '__main__':
    if '--liste' in sys.argv:
        liste()
    else:
        arsivle()
        liste()
