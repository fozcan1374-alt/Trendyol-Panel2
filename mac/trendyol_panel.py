#!/usr/bin/env python3
"""TRENDYOL PANELİ — masaüstü arayüz (pywebview).
Veri: ~/Desktop/Trendyol_Karlilik.xlsx (uygulama okur ve Maliyetler'e YAZAR).
Başlat: 'Trendyol Paneli.command' çift tık."""
import openpyxl, json, math, warnings, subprocess, sys, datetime, os
from openpyxl.styles import Font, PatternFill, Alignment
warnings.filterwarnings('ignore')
import webview
import ayar          # yollar paketin konumundan türetilir

XLSX=ayar.XLSX
TARIFE_ARSIV=ayar.TARIFE_ARSIV   # tarife_arsivle.py buraya yazar
KDV=0.20        # stopaj matrahı = ciro / (1+KDV); ürün bazlı farklı KDV gerekirse burada esnetilir
STOPAJ=0.01     # e-ticaret stopajı, KDV hariç tutar üzerinden
HB_TAHSILAT=0.008  # HB tahsilat bedeli: ciro × %0,8 (panel ekstresiyle birebir doğrulandı)
HB_KOM_KDV=1.20    # API komisyonu KDV HARİÇ gelir; panelde KDV dahil kesiliyor
HB_HIZMET=12.60    # HB hizmet bedeli — sipariş başına sabit (panel ekstresinden)
TOLERANS=0.01      # buybox'ı almak için min kârın kaç PUAN altına inilebilir (0,01 = 1 puan).
                   # Gerekçe: satmayan yüksek fiyatın kârı sıfırdır; 1 puan feragat edip
                   # buybox'ı almak daha iyidir. Sıfırlamak = "min kâr kutsal" davranışı.
TAVAN_KAR=0.20     # Hedefler sekmesinde "Tavan Kâr %" boşsa geçerli üst sınır (kullanıcı:
                   # elektronikte %20). Kâr marjı bu tavana ulaşmışsa "fiyat yükselt" denmez —
                   # buybox bizdeyken üstteki satıcı çok yukarıdaysa bile.
PLATFORM=13.19     # Trendyol Platform Hizmet Bedeli — GÖNDERİ başına sabit, KDV dahil (10,99+KDV).
                   # 180 günde 51 kaydın 50'si bu birime tam bölünüyor; toplu kesiliyor, ~1 ay gecikmeli.

def nb(x):
    """Barkod normalizasyonu — HB barkodları başında sıfır dolgulu geliyor (0000064787788),
    Excel/Trendyol tarafı sıfırsız (64787788). Tüm eşleşmeler bu sadeleştirilmiş hâl üzerinden."""
    t=str(x or '').strip()
    k=t.lstrip('0')
    return k or t

def _f(v):
    try:
        if v is None or v=='': return None
        return float(v)
    except Exception: return None

class Veri:
    def __init__(self):
        self.donem=1
        self.yukle()

    def yukle(self):
        wb=openpyxl.load_workbook(XLSX, data_only=True)
        kt=wb['Komisyon Tarifeleri']
        # kargo
        kg=wb['Kargo']
        self.kargo_desi={int(kg.cell(r,1).value): float(kg.cell(r,2).value)
                         for r in range(2,80) if kg.cell(r,1).value is not None}
        self.desi_max=max(self.kargo_desi) if self.kargo_desi else 20
        self.kargo_dusuk=(float(kg['E2'].value or 47.4), float(kg['E3'].value or 85.4))
        self.hb_kargo={int(kg.cell(r,1).value): float(kg.cell(r,3).value)
                       for r in range(2,30) if kg.cell(r,1).value is not None and kg.cell(r,3).value}
        # hedefler
        hd=wb['Hedefler']
        self.hedefler={}
        for r in range(2, hd.max_row+1):
            k=hd.cell(r,1).value
            if k: self.hedefler[str(k)]=( _f(hd.cell(r,2).value) or 0.04, _f(hd.cell(r,3).value) or 0.10,
                                          _f(hd.cell(r,4).value) or TAVAN_KAR )   # D = tavan kâr %
        # maliyetler
        ml=wb['Maliyetler']
        self.girdi={}; self.satirlar=[]
        for r in range(2, ml.max_row+1):
            bk=ml.cell(r,1).value
            if bk:
                _kayit=dict(
                    maliyet=_f(ml.cell(r,3).value), desi=_f(ml.cell(r,4).value),
                    buybox=_f(ml.cell(r,5).value), canli=_f(ml.cell(r,6).value),
                    stok=ml.cell(r,7).value, sync=ml.cell(r,8).value,
                    kom_manuel=_f(ml.cell(r,10).value), buybox2=_f(ml.cell(r,12).value),
                    ad=str(ml.cell(r,2).value or ''),
                    kategori=str(ml.cell(r,13).value or ''),   # M — Ürünler ekranı yazar
                    api_kom=_f(ml.cell(r,14).value),           # N — Trendyol API komisyonu
                    bb_sira=_f(ml.cell(r,15).value),           # O — buybox sırası (1 = bizde)
                    bb_coklu=(None if ml.cell(r,16).value in (None,'') else bool(ml.cell(r,16).value)))  # P
                self.girdi[str(bk).strip()]=_kayit
                self.girdi[nb(bk)]=_kayit
                self.satirlar.append((nb(bk), _kayit))   # tarife dışı ürünler de listelenebilsin
        # plus
        self.plus={}
        if 'Plus Tarifeleri' in wb.sheetnames:
            pl=wb['Plus Tarifeleri']
            for r in range(2, pl.max_row+1):
                bk=pl.cell(r,3).value
                if bk:
                    self.plus[str(bk).strip()]=dict(limit=_f(pl.cell(r,13).value),
                        kom1=_f(pl.cell(r,15).value), kom2=_f(pl.cell(r,17).value))
        # tarifeler
        self.donem_tarih=[str(kt.cell(2,15).value or ''), str(kt.cell(2,20).value or '')]
        self.urunler=[]
        for r in range(2, kt.max_row+1):
            bk=kt.cell(r,2).value
            if bk is None: continue
            self.urunler.append(dict(
                ad=str(kt.cell(r,1).value or ''), barkod=str(bk).strip(),
                kategori=str(kt.cell(r,6).value or ''),
                limit=[_f(kt.cell(r,c).value) for c in (9,10,11,12,13,14)],  # I J K L M N
                kom1=[_f(kt.cell(r,c).value) for c in (16,17,18,19)],        # P Q R S
                kom2=[_f(kt.cell(r,c).value) for c in (21,22,23,24)],        # U V W X
                guncel_kom=_f(kt.cell(r,26).value),                          # Z
                tsf=_f(kt.cell(r,27).value)))                                # AA

    # ---- hesap
    def kargo(self, fiyat, desi):
        if fiyat is None: return None
        if fiyat<=199.99: return self.kargo_dusuk[0]
        if fiyat<=399.99: return self.kargo_dusuk[1]
        d=int(max(1,min(self.desi_max, round(desi or 1))))
        return self.kargo_desi.get(d, 99.46)

    STOPAJ_ORAN=STOPAJ/(1+KDV)      # fiyat üzerinden efektif stopaj = %1 / 1,20 = %0,8333

    def kar(self, fiyat, kom, maliyet, desi):
        """Net kâr = fiyat − komisyon − stopaj − kargo − platform hizmet bedeli − ürün maliyeti.
        Ciro raporundaki kalemlerin birebir aynısı."""
        if None in (fiyat,kom,maliyet): return None
        return (fiyat*(1-kom/100) - fiyat*self.STOPAJ_ORAN
                - maliyet - self.kargo(fiyat,desi) - PLATFORM)

    @property
    def bant(self):
        """barkod → komisyon tarifesi satırı (fırsat ekranı barem hesabı için)."""
        return {u['barkod']:u for u in self.urunler}

    def kom_bant(self, limit, kom, fiyat):
        """Fiyatın düştüğü baremin komisyonu (tarifeli ürün)."""
        if fiyat is None: return None
        I,J,K,L,M,N=limit
        if I is not None and fiyat>=I: return kom[0]
        if K is not None and fiyat>=K: return kom[1]
        if M is not None and fiyat>=M: return kom[2]
        return kom[3]

    BB_FARK=0.01     # buybox bizdeyken 2. satıcıyla aramızdaki AZAMİ fark (kullanıcı kuralı,
                     # 2026-09-01): daha fazlaysa fiyat gereksiz düşük kalmış demektir.

    def hesapla(self):
        """BUYBOX + FİYAT KONTROL listesi — STOKTA olan TÜM ürünler, tarife dosyasında
        olsun olmasın (kullanıcı kuralı 2026-09-01).

        Tarifesiz üründe komisyon TEK ORAN (manuel J → API N) ile hesaplanır; barem kaskadı
        yapılamaz, çünkü hangi fiyatta oranın düştüğü bilinmiyor — satırda rozetle belirtilir.

        Buybox durumları:
          👤 tek satıcı        → aksiyon yok, sadece rozet
          🏆 buybox bizde      → 2. satıcıyla fark ≤ %1, fiyat doğru yerde
          💸 fiyat düşük kalmış→ fark > %1, hedef = 2. fiyat × 0,99'a YÜKSELT
          ⚔️ buybox rakipte    → mevcut öneri motoru (buybox−1 / barem düşür)
        Sıralama: önce aksiyon gerektirenler (kaçan kâr büyükten küçüğe), sonra gerisi."""
        out=[]
        tf={u['barkod']:u for u in self.urunler}
        for bk,g in self.satirlar:
            if not ((_f(g.get('stok')) or 0)>0): continue
            t=tf.get(bk)
            m=g.get('maliyet'); desi=g.get('desi') or 1
            bb=g.get('buybox'); bb2=g.get('buybox2'); canli=g.get('canli')
            sira=g.get('bb_sira'); coklu=g.get('bb_coklu')
            ad=(t or {}).get('ad') or g.get('ad') or bk
            kategori=(t or {}).get('kategori') or g.get('kategori') or ''
            minp,hedp,tavan=self.hedefler.get(kategori,(0.04,0.10,TAVAN_KAR))
            kd=self.kargo_desi.get(int(max(1,min(self.desi_max,round(desi)))),99.46)
            if t:
                tarifesiz=False; kom_tek=None
                kom=t['kom1'] if self.donem==1 else t['kom2']
                I,J,K,L,M,N=t['limit']
                tsf=t['tsf'] or I
                # bantlar: [alt, üst(None=∞), kom, temsili fiyat]
                bantlar=[dict(alt=I, ust=None, kom=kom[0], fiyat=max(tsf or 0, I or 0)),
                         dict(alt=K, ust=J, kom=kom[1], fiyat=J),
                         dict(alt=M, ust=L, kom=kom[2], fiyat=L),
                         dict(alt=0, ust=N, kom=kom[3], fiyat=N)]
                komf=lambda f,_l=t['limit'],_k=kom: self.kom_bant(_l,_k,f)
                sirali=[(0,N,kom[3]),(M,L,kom[2]),(K,J,kom[1]),(I,None,kom[0])]
            else:
                tarifesiz=True
                kom_tek=g.get('kom_manuel') if g.get('kom_manuel') is not None else g.get('api_kom')
                bantlar=[dict(alt=None, ust=None, kom=kom_tek, fiyat=canli)]
                komf=lambda f,_k=kom_tek: _k
                sirali=[(0,None,kom_tek)]
            for b in bantlar:
                b['kar']=self.kar(b['fiyat'], b['kom'], m, desi)
                b['pct']=(b['kar']/b['fiyat']) if (b['kar'] is not None and b['fiyat']) else None
            def kaskad(p):
                """p kârını koruyan en düşük fiyat (tarifesizde tek bant)."""
                if m is None: return None
                for alt,ust,kk in sirali:
                    if kk is None: continue
                    pay=1-kk/100-self.STOPAJ_ORAN-p
                    if pay<=0: continue
                    fmin=math.ceil((m+kd+PLATFORM)/pay)
                    cand=max(alt or 0, fmin)
                    if ust is None or cand<=ust: return cand
                return None
            # ---- buybox altı senaryosu
            bb_f=bb-1 if bb else None
            bb_kom=komf(bb_f) if bb_f is not None else None
            bb_kar=self.kar(bb_f, bb_kom, m, desi) if bb_f is not None else None
            bb_pct=(bb_kar/bb_f) if (bb_kar is not None and bb_f) else None
            # Kâr% bir bandın İÇİNDE fiyatla birlikte hep artar → her bandın en kârlı noktası
            # daima o bandın ÜST sınırıdır. Buybox'ın altında kalmak şartıyla en iyisini bul.
            en_iyi=None
            if bb_f is not None and m is not None:
                for bi,b in enumerate(bantlar):
                    ust=b['ust'] if b['ust'] is not None else bb_f
                    f_=math.floor(min(ust, bb_f))
                    if b['kom'] is None or f_<(b['alt'] or 0) or f_<=0: continue
                    k_=self.kar(f_, b['kom'], m, desi)
                    if k_ is None: continue
                    p_=k_/f_
                    if en_iyi is None or p_>en_iyi['pct']:
                        en_iyi=dict(fiyat=f_, kom=b['kom'], kar=k_, pct=p_, bant=bi)
            if m is None:
                oneri=oneri_kar=oneri_pct=None; durum='MALİYET GİRİLMEMİŞ'
            elif bb_f is not None and bb_pct is not None and bb_pct>=minp:
                oneri=bb_f; oneri_kar=bb_kar; oneri_pct=bb_pct; durum='✓ BUYBOX ALTINA GİR'
            elif en_iyi is not None and en_iyi['pct']>=minp:
                oneri=en_iyi['fiyat']; oneri_kar=en_iyi['kar']; oneri_pct=en_iyi['pct']
                durum='✓ BAREM DÜŞÜR → BUYBOX'
            elif en_iyi is not None and en_iyi['pct']>=minp-TOLERANS:
                oneri=en_iyi['fiyat']; oneri_kar=en_iyi['kar']; oneri_pct=en_iyi['pct']
                durum='✓ BAREM DÜŞÜR → BUYBOX (tolerans)'
            elif bb_f is not None:
                oneri=kaskad(minp); durum='BUYBOX ALTI KÂRSIZ → min kârlı fiyat'
            else:
                oneri=kaskad(hedp); durum='BUYBOX YOK → hedef kârla min fiyat'
            if m is not None and durum.startswith(('BUYBOX ALTI','BUYBOX YOK')) and oneri:
                oneri_kar=self.kar(oneri, komf(oneri), m, desi)
                oneri_pct=(oneri_kar/oneri) if oneri else None
            def fiyat_tavan(p, baslangic):
                """Kâr marjı tam olarak p olan fiyat. Komisyon ve kargo fiyata bağlı olduğu için
                sabit nokta iterasyonu; bu fiyatın ÜSTÜ tavanı aşar."""
                if m is None or p is None: return None
                f=baslangic or 100
                for _ in range(5):
                    kk=komf(f)
                    if kk is None: return None
                    pay=1-kk/100-self.STOPAJ_ORAN-p
                    if pay<=0: return None
                    yeni_f=(m+self.kargo(f,desi)+PLATFORM)/pay
                    if abs(yeni_f-f)<0.5: f=yeni_f; break
                    f=yeni_f
                return f
            # ---- BUYBOX KONTROLÜ — tarifeden bağımsız, her stoklu ürüne uygulanır
            fark=hedef=kacan=None; aksiyon=False; kisit=None; tavan_fiyat=None
            # mevcut fiyattaki marj — kabul aralığı kontrolünün ölçüsü
            kar_simdi=self.kar(canli, komf(canli), m, desi) if (m is not None and canli) else None
            canli_pct=(kar_simdi/canli) if (kar_simdi is not None and canli) else None
            if coklu is False:
                bb_durum='tek'; durum='👤 TEK SATICI'; grup=5
                oneri=oneri_kar=oneri_pct=None
            elif coklu and sira and int(sira)==1:
                # KÂR KABUL ARALIĞI (kullanıcı kuralı 2026-09-01): marj [min, tavan] arasındaysa
                # fiyat DOĞRUDUR — üstteki satıcı ne kadar yukarıda olursa olsun uyarı verilmez.
                # Mobilya %25–40, elektronik ... –%20. Uyarı yalnız marj MİN'in altındayken.
                fark=((bb2-canli)/canli) if (bb2 and canli) else None
                if canli_pct is None:
                    durum='MALİYET GİRİLMEMİŞ'; bb_durum='maliyetsiz'; grup=3
                    oneri=oneri_kar=oneri_pct=None
                elif canli_pct<minp:
                    # hedef kâra çek; ama 2. satıcının %1 altını ve tavanı aşma
                    tavan_fiyat=fiyat_tavan(tavan, canli)
                    f_hed=fiyat_tavan(min(hedp,tavan), canli)
                    hedef=math.floor(f_hed) if f_hed else None; kisit='hedef'
                    ust=math.floor(bb2*(1-self.BB_FARK)) if bb2 else None
                    if ust is None: kisit='ust_bilinmiyor'      # 2. satıcı fiyatı gelmedi
                    elif hedef is not None and ust<hedef:
                        hedef=ust; kisit='rakip'
                    if hedef is not None and hedef>canli:
                        k_yeni=self.kar(hedef, komf(hedef), m, desi)
                        kacan=(k_yeni-kar_simdi) if k_yeni is not None else (hedef-canli)
                        oneri=hedef; oneri_kar=k_yeni
                        oneri_pct=(k_yeni/hedef) if (k_yeni is not None and hedef) else None
                        durum='💸 KÂR DÜŞÜK — FİYAT YÜKSELT'; bb_durum='dusuk'; grup=1; aksiyon=True
                    else:
                        # marj min'in altında ama rakip yer bırakmıyor: yükseltirsek buybox gider
                        durum='⚠️ KÂR DÜŞÜK · rakip yer bırakmıyor'; bb_durum='sikisik'; grup=3
                        oneri=oneri_kar=oneri_pct=None; hedef=None
                elif canli_pct>tavan:
                    durum='🏆 BUYBOX BİZDE · kâr tavanın üstünde'; bb_durum='ustunde'; grup=4
                    oneri=oneri_kar=oneri_pct=None
                else:
                    durum='🏆 BUYBOX BİZDE'; bb_durum='bizde'; grup=4
                    oneri=oneri_kar=oneri_pct=None
            elif coklu and (sira is None or int(sira)>1):
                # DİKKAT: buybox sırf fiyatla belirlenmiyor (satıcı puanı, teslimat hızı, stok).
                # Fiyatımız buybox'tan ZATEN düşükken "buybox altına gir" demek saçma —
                # o öneri fiyatı YÜKSELTİR ve buybox'ı yine almaz. (Pad X9A: biz 10.733,
                # buybox 10.735, sıra 2 → eski motor 10.734 öneriyordu.)
                if canli is not None and bb is not None and canli<bb:
                    durum='⚔️ RAKİPTE · fiyatımız zaten daha düşük'; bb_durum='rakipte_ucuz'; grup=3
                    oneri=oneri_kar=oneri_pct=None
                else:
                    bb_durum='rakipte'
                    aksiyon=(oneri is not None and bb_f is not None and oneri<=bb_f)
                    grup=2 if aksiyon else 3
            else:
                bb_durum='veriyok'; durum='BUYBOX VERİSİ YOK'; grup=3
                oneri=oneri_kar=oneri_pct=None
            if m is None and bb_durum in ('rakipte','veriyok'): grup=3
            # plus
            p=self.plus.get(bk)
            plus=None
            if p and p['limit']:
                pk=p['kom1'] if self.donem==1 else p['kom2']
                pkar=self.kar(p['limit'], pk, m, desi)
                plus=dict(limit=p['limit'], kom=pk, kar=pkar,
                          pct=(pkar/p['limit']) if (pkar is not None and p['limit']) else None,
                          fark=(pkar-oneri_kar) if (pkar is not None and oneri_kar is not None) else None)
            # simülatör aralığı (tarifesizde bant sınırı yok, canlı fiyattan türetilir)
            if tarifesiz:
                taban=canli or bb or 100
                sim_lo=max(1, math.floor(taban*0.6)); sim_hi=math.ceil(max(taban, bb2 or 0, bb or 0)*1.25)
            else:
                sim_lo=max(1, math.floor((bantlar[3]['ust'] or 100)*0.72))
                sim_hi=math.ceil(max(bantlar[0]['fiyat'] or 0, bb or 0)*1.06)
            out.append(dict(ad=ad, barkod=bk, kategori=kategori,
                maliyet=m, desi=desi, buybox=bb, buybox2=bb2,
                canli=canli, stok=g.get('stok'),
                sync=str(g.get('sync') or ''), bantlar=bantlar, minp=minp, hedp=hedp,
                bb_f=bb_f, bb_kom=bb_kom, bb_kar=bb_kar, bb_pct=bb_pct,
                oneri=oneri, oneri_kar=oneri_kar, oneri_pct=oneri_pct, durum=durum,
                en_iyi=en_iyi, minp_tol=minp-TOLERANS,
                satilabilir=(oneri is not None and bb_f is not None and oneri<=bb_f),
                tarifesiz=tarifesiz, kom_tek=kom_tek,
                bb_durum=bb_durum, bb_sira=(int(sira) if sira else None), bb_coklu=coklu,
                fark=fark, hedef=hedef, kacan=kacan, aksiyon=aksiyon, grup=grup,
                tavan=tavan, kisit=kisit, tavan_fiyat=tavan_fiyat, canli_pct=canli_pct,
                sim_lo=sim_lo, sim_hi=sim_hi,
                plus=plus, kargo_desi_tl=self.kargo_desi.get(int(max(1,min(self.desi_max,round(desi)))),99.46)))
        # önce aksiyon gerekenler (kaçan kâr / öneri kârı büyükten küçüğe), sonra gerisi
        out.sort(key=lambda x:(x['grup'], -((x['kacan'] if x['kacan'] is not None else (x['oneri_kar'] or 0)) or 0)))
        return out

class Api:
    def __init__(self): self.v=Veri()
    def durum(self):
        return json.dumps(dict(donem=self.v.donem, tarih=self.v.donem_tarih,
                               urunler=self.v.hesapla()), ensure_ascii=False)
    def donem_sec(self, d):
        self.v.donem=int(d); return self.durum()
    def kaydet(self, barkod, maliyet, desi, buybox):
        wb=openpyxl.load_workbook(XLSX)
        ml=wb['Maliyetler']
        satir=None
        for r in range(2, ml.max_row+1):
            if str(ml.cell(r,1).value or '').strip()==str(barkod):
                satir=r; break
        if satir is None:
            satir=ml.max_row+1; ml.cell(satir,1,str(barkod))
        def yaz(c,v):
            ml.cell(satir,c, float(v) if v not in (None,'','null') else None)
        yaz(3,maliyet); yaz(4,desi); yaz(5,buybox)
        wb.save(XLSX)
        self.v.yukle(); return self.durum()
    def fiyat_gonder(self, barkod, fiyat):
        """Önerilen fiyatı Trendyol'a bas (SADECE kullanıcı onayıyla çağrılır)."""
        import requests, base64, time as _t
        try:
            fiyat=float(fiyat)
            cfg=json.load(open(ayar.TY_API))
            tok=base64.b64encode(f"{cfg['api_key']}:{cfg['api_secret']}".encode()).decode()
            H={'Authorization':f'Basic {tok}','User-Agent':f"{cfg['seller_id']} - SelfIntegration",'Content-Type':'application/json'}
            # mevcut listPrice + stok bul
            lp=q=None
            for page in range(10):
                r=requests.get(f"https://apigw.trendyol.com/integration/product/sellers/{cfg['seller_id']}/products/approved/inventory-and-price",
                               params={'page':page,'size':100},headers=H,timeout=30).json()
                for pr in r.get('content',[]):
                    for v in pr.get('variants',[]):
                        if str(v.get('barcode','')).strip()==str(barkod):
                            lp=v.get('listPrice'); q=v.get('quantity'); break
                if lp is not None or page>=r.get('totalPages',1)-1: break
            if q is None:
                return json.dumps({'ok':False,'mesaj':'Ürün API listesinde bulunamadı'},ensure_ascii=False)
            item={'barcode':str(barkod),'salePrice':fiyat,'listPrice':max(fiyat, lp or 0),'quantity':int(q)}
            rr=requests.post(f"https://apigw.trendyol.com/integration/inventory/sellers/{cfg['seller_id']}/products/price-and-inventory",
                             json={'items':[item]},headers=H,timeout=30)
            if rr.status_code!=200:
                return json.dumps({'ok':False,'mesaj':f'HTTP {rr.status_code}: {rr.text[:200]}'},ensure_ascii=False)
            bid=rr.json().get('batchRequestId')
            _t.sleep(3)
            b=requests.get(f"https://apigw.trendyol.com/integration/product/sellers/{cfg['seller_id']}/products/batch-requests/{bid}",headers=H,timeout=20).json()
            hata=[str(i.get('failureReasons')) for i in b.get('items',[]) if i.get('status')=='FAILED']
            if hata:
                return json.dumps({'ok':False,'mesaj':'Trendyol reddetti: '+'; '.join(hata)[:300]},ensure_ascii=False)
            return json.dumps({'ok':True,'mesaj':f'Gönderildi ✓ (batch {bid[:13]}...). Yeni fiyat birkaç dk içinde yayına girer; sonra Güncelle ile doğrula.','item':item},ensure_ascii=False)
        except Exception as e:
            return json.dumps({'ok':False,'mesaj':f'Hata: {e}'},ensure_ascii=False)

    def _api_komisyon(self):
        """Barkod → güncel komisyon oranı, ürün servisinin varyant `commission` alanından.
        Bu alan dokümanda GEÇMİYOR ama %100 dolu ve haftalık tarife dosyasıyla birebir aynı
        (14 ortak üründe 14/14). Ürünün O ANKİ fiyatına karşılık gelen oran — barem tablosunun
        tamamını vermez, o yüzden tarife Excel'i yine gerekli."""
        if getattr(self,'_apikom',None) is not None: return self._apikom
        import requests, base64
        cfg=json.load(open(ayar.TY_API)); sid=cfg['seller_id']
        tok=base64.b64encode(f"{cfg['api_key']}:{cfg['api_secret']}".encode()).decode()
        H={'Authorization':f'Basic {tok}','User-Agent':f'{sid} - SelfIntegration'}
        out={}
        try:
            for page in range(30):
                r=requests.get(f'https://apigw.trendyol.com/integration/product/sellers/{sid}/products/approved',
                               params={'page':page,'size':100},headers=H,timeout=30).json()
                for p_ in r.get('content',[]):
                    for va in p_.get('variants',[]):
                        b=str(va.get('barcode','')).strip()
                        if b and va.get('commission') is not None: out[b]=float(va['commission'])
                if page>=r.get('totalPages',1)-1: break
        except Exception: pass
        self._apikom=out; return out

    def _desi_kalibre(self, kalemler, sip):
        """Kargo faturasındaki GERÇEK desiyi ürüne yazar. Trendyol paketin desisini ölçtüğü için
        (kutu+dolgu dahil) elle girilen desi düşük kalıyor. Sadece tek ürünlü gönderiler
        kullanılır — çok ürünlü pakette desi ürüne pay edilemez. Her ciro raporunda arka planda
        çalışır, yeni fatura geldikçe desi kendiliğinden güncellenir.
        Ölçümler YEREL DEFTERDE birikir (`_kargo_desi_gozlem.json`): rapor 7 günlük mü 60 günlük mü
        diye medyan değişmesin, gördüğümüz her gönderi kalıcı olsun diye. Defter olmadan pencere
        daraldıkça medyan oynuyor ve desi her raporda ileri geri yazılıyordu."""
        import statistics as _st, os
        DEFTER=ayar.DESI_DEFTER
        try: defter=json.load(open(DEFTER))
        except Exception: defter={}
        yeni_gozlem=0
        for c in kalemler:
            if str(c.get('shipmentPackageType'))!='Gönderi Kargo Bedeli': continue
            ds=c.get('desi')
            if not ds: continue
            onum=str(c.get('orderNumber')); o=sip.get(onum)
            if not o: continue
            bks={x[0] for x in o['ln'] if x[0]}
            if len(bks)!=1: continue                      # çok ürünlü paket → pay edilemez
            bk=bks.pop(); q=sum(x[2] for x in o['ln']) or 1
            anahtar=f"{onum}:{c.get('parcelUniqueId') or ''}"
            if anahtar not in defter:
                defter[anahtar]=dict(barkod=bk, desi=float(ds), adet=q); yeni_gozlem+=1
        if yeni_gozlem:
            try: json.dump(defter, open(DEFTER,'w'), ensure_ascii=False)
            except Exception: pass
        gozlem={}
        for k in defter.values():
            gozlem.setdefault(k['barkod'],[]).append((k['desi']/max(1,k.get('adet') or 1), k.get('adet') or 1))
        if not gozlem: return dict(gozlem=0, yazilan=0, yeni=0, urun=0, satirlar=[])
        wb=openpyxl.load_workbook(XLSX)
        ml=wb['Maliyetler']
        if ml.cell(1,11).value!='Desi Kaynağı':
            c=ml.cell(1,11,'Desi Kaynağı'); c.font=Font(bold=True,color='FFFFFF',size=10)
            c.fill=PatternFill('solid',fgColor='1F4E78'); c.alignment=Alignment(horizontal='center',wrap_text=True)
            ml.column_dimensions['K'].width=26
        satir={}
        for r in range(2, ml.max_row+1):
            b=ml.cell(r,1).value
            if b: satir[str(b).strip()]=r; satir.setdefault(nb(b), r)
        bugun=datetime.datetime.now().strftime('%d.%m.%Y')
        yaz=0; yeni=0; sat=[]
        for bk,ob in gozlem.items():
            tek=[d_ for d_,q in ob if q==1] or [d_ for d_,q in ob]   # tercih: tek adetli gönderi
            ger=round(_st.median(tek),1)
            r=satir.get(bk) or satir.get(nb(bk))
            if r is None:
                r=ml.max_row+1; ml.cell(r,1,bk); satir[bk]=r
            eski=ml.cell(r,4).value
            try: eski=float(eski) if eski is not None else None
            except Exception: eski=None
            if eski is None or abs(eski-ger)>0.4:
                ml.cell(r,4,ger); ml.cell(r,11,f'kargo faturası · {len(tek)} gönderi · {bugun}')
                sat.append(dict(barkod=bk, ad=str(ml.cell(r,2).value or '')[:44], eski=eski, yeni=ger, n=len(tek)))
                yaz+=1
                if eski is None: yeni+=1
        if yaz:
            wb.save(XLSX)
            self.v.yukle()          # tahminler güncel desiyle yapılsın
        sat.sort(key=lambda x:-(x['yeni']-(x['eski'] or 0)))
        return dict(gozlem=sum(len(v) for v in gozlem.values()), yazilan=yaz, yeni=yeni,
                    urun=len(gozlem), satirlar=sat[:30])

    def ciro(self, gun):
        """Ürün bazlı ciro raporu — SİPARİŞ TARİHİ bazlı (Trendyol panel raporuyla aynı taban).
        Komisyon: hakediş kaydı oluşanlarda gerçek, oluşmayanlarda güncel orandan TAHMİN (≈).
        Kargo: fatura kalemi sipariş tarihine eşlenir; faturasızlara desi tarifesinden tahmin."""
        import requests, base64, time as _t
        gun=int(gun)
        cfg=json.load(open(ayar.TY_API))
        sid=cfg['seller_id']
        tok=base64.b64encode(f"{cfg['api_key']}:{cfg['api_secret']}".encode()).decode()
        H={'Authorization':f'Basic {tok}','User-Agent':f'{sid} - SelfIntegration'}
        A='https://apigw.trendyol.com/integration'
        simdi=_t.time(); donem_bas=(simdi-gun*86400)*1000
        def parcala(toplam_gun):
            out=[]; kalan=toplam_gun; uc=simdi
            while kalan>0:
                al=min(13,kalan); out.append((uc-al*86400, uc)); uc-=al*86400; kalan-=al
            return out
        def sayfali(url, params, size=500):
            page=0; out=[]
            while True:
                pr=dict(params); pr['page']=page; pr['size']=size
                r=requests.get(url,params=pr,headers=H,timeout=30).json()
                out+=r.get('content',[])
                if page>=r.get('totalPages',1)-1: break
                page+=1
            return out
        urun={}; guncel_kom={u['barkod']:(u.get('guncel_kom') or 0) for u in self.v.urunler}
        for _b,_k in self._api_komisyon().items():      # tarifede olmayanlar için API oranı
            if not guncel_kom.get(_b): guncel_kom[_b]=_k
        def d(bk):
            return urun.setdefault(bk,{'ad':'','ciro':0.0,'kom':0.0,'kom_t':0.0,'kom_o':0.0,
                                      'kargo':0.0,'kargo_t':0.0,'platform':0.0,'adet':0})
        # not: stopaj ve maliyet aşağıda satır bazında eklenir
        # --- siparişler (gun+45 geriye: kargo/hakediş eşleşmesi için harita geniş tutulur)
        sip={}
        for s_,e_ in parcala(gun+45):
            for o in sayfali(f'{A}/order/sellers/{sid}/orders',
                             {'startDate':int(s_*1000),'endDate':int(e_*1000)},200):
                ln=[]; komor={}
                for l in o.get('lines',[]):
                    bk=str(l.get('barcode','')).strip()
                    q=l.get('quantity') or 1
                    tut=(l.get('price') or 0)*q
                    ln.append((bk,tut,q))
                    # sipariş satırındaki `commission` = SATIŞ ANINDAKİ komisyon ORANI.
                    # Hakedişle birebir doğrulandı (%22 ↔ %22,0) ve hakediş beklemeden gelir.
                    if bk and l.get('commission') is not None:
                        try: komor[bk]=float(l['commission'])
                        except Exception: pass
                    if bk and not d(bk)['ad']: d(bk)['ad']=str(l.get('productName') or '')[:60]
                sip[str(o.get('orderNumber'))]={'ln':ln,'tarih':o.get('orderDate') or 0,
                                               'durum':str(o.get('status') or ''),'komor':komor}
        # --- CİRO: dönem içi, iptal olmayan siparişler (sipariş tarihi bazlı)
        for onum,o in sip.items():
            if o['tarih']<donem_bas or o['durum']=='Cancelled': continue
            for bk,tut,q in o['ln']:
                d(bk)['ciro']+=tut; d(bk)['adet']+=q
        # --- KOMİSYON: hakedişten gerçek (sipariş dönem içiyse); iade düşümü
        settled=set()
        for s_,e_ in parcala(gun+45):
            for tt in ('Sale','Return'):
                for k in sayfali(f'{A}/finance/che/sellers/{sid}/settlements',
                                 {'startDate':int(s_*1000),'endDate':int(e_*1000),'transactionType':tt}):
                    onum=str(k.get('orderNumber')); o=sip.get(onum)
                    if not o or o['tarih']<donem_bas: continue
                    bk=str(k.get('barcode','')).strip(); x=d(bk)
                    if tt=='Sale':
                        x['kom']+=(k.get('commissionAmount') or 0); settled.add((onum,bk))
                    else:
                        brut=(k.get('sellerRevenue') or 0)+(k.get('commissionAmount') or 0)
                        x['ciro']-=brut; x['adet']-=1; x['kom']-=(k.get('commissionAmount') or 0)
        # komisyon tahmini: hakedişi henüz oluşmamış dönem içi satışlar
        for onum,o in sip.items():
            if o['tarih']<donem_bas or o['durum']=='Cancelled': continue
            for bk,tut,q in o['ln']:
                if (onum,bk) in settled: continue
                sip_oran=o.get('komor',{}).get(bk)
                oran=sip_oran or guncel_kom.get(bk) or 15.0
                tl_=tut*oran/100
                d(bk)['kom']+=tl_
                # sipariş satırından gelen oran GERÇEK (sadece hakediş kaydı henüz yok);
                # gerçekten tahmin olan yalnız oranı hiçbir yerden bulunamayanlar
                if sip_oran: d(bk)['kom_o']+=tl_
                else: d(bk)['kom_t']+=tl_
        # --- KARGO: fatura kalemleri sipariş tarihine eşlenir; faturasızlara tahmin
        kalemler=[]
        for s_,e_ in parcala(gun+45):
            for k in sayfali(f'{A}/finance/che/sellers/{sid}/otherfinancials',
                             {'startDate':int(s_*1000),'endDate':int(e_*1000),'transactionType':'DeductionInvoices'}):
                if 'Kargo' not in str(k.get('transactionType','')): continue
                seri=k.get('commissionInvoiceSerialNumber') or k.get('id')
                try:
                    it=requests.get(f'{A}/finance/che/sellers/{sid}/cargo-invoice/{seri}/items',
                                    params={'page':0,'size':1000},headers=H,timeout=30).json()
                    kalemler+=(it if isinstance(it,list) else it.get('content',[]))
                except Exception: pass
        # faturadaki GERÇEK desiyi ürünlere yaz (tahmin bu güncel desiyle yapılsın diye önce çalışır)
        self._desi_bilgi=self._desi_kalibre(kalemler, sip)
        faturali=set()
        for c in kalemler:
            onum=str(c.get('orderNumber')); tut=c.get('amount') or 0
            o=sip.get(onum); faturali.add(onum)
            if o and o['tarih']>=donem_bas:
                top=sum(x[1] for x in o['ln']) or 1
                for bk,pay,_ in o['ln']: d(bk)['kargo']+=tut*(pay/top)
        for onum,o in sip.items():
            if onum in faturali or o['tarih']<donem_bas: continue
            # kargo faturası henüz gelmemiş her sipariş tahminlenir (Created/Picking dahil —
            # bunlar da sevk edilecek); iptal/iade dışarıda kalır
            if o['durum'] in ('Cancelled','UnDelivered','Returned'): continue
            for bk,pay,_ in o['ln']:
                desi_=self.v.girdi.get(bk,{}).get('desi') or 1
                th=self.v.kargo_desi.get(int(max(1,min(self.v.desi_max,round(desi_)))),99.46)
                d(bk)['kargo']+=th; d(bk)['kargo_t']+=th
        # --- PLATFORM HİZMET BEDELİ: gönderi başına sabit, kalemlere ciro payıyla dağıtılır
        for onum,o in sip.items():
            if o['tarih']<donem_bas or o['durum']=='Cancelled': continue
            top=sum(x[1] for x in o['ln']) or 1
            for bk,pay,_ in o['ln']: d(bk)['platform']+=PLATFORM*(pay/top)
        for u in self.v.urunler:
            if u['barkod'] in urun and not urun[u['barkod']]['ad']:
                urun[u['barkod']]['ad']=u['ad'][:60]
        self._ham=dict(gun=gun, urun=urun)
        return self._ciro_sonuc()

    def _ciro_sonuc(self):
        """Cache'lenmiş ham veriden (ciro/kom/kargo/adet) maliyet+stopaj+net'i yeniden hesaplar.
        Maliyet değişince API'yi tekrar taramaya gerek kalmaz."""
        if not getattr(self,'_ham',None): return json.dumps(dict(gun=0,satirlar=[],eksik=[],toplam={}),ensure_ascii=False)
        gun=self._ham['gun']; urun=self._ham['urun']
        satirlar=[]
        for bk,v in urun.items():
            if not (abs(v['ciro'])>0.01 or v['kargo']>0.01): continue
            v.setdefault('platform',0.0); v.setdefault('kom_o',0.0)
            stopaj=(v['ciro']/(1+KDV))*STOPAJ
            mb=self.v.girdi.get(bk,{}).get('maliyet')
            mal=(mb*v['adet']) if (mb is not None and v['adet']>0) else (0.0 if mb is not None else None)
            kom_all=v['kom']+stopaj                      # stopaj komisyona yedirildi
            net=(v['ciro']-kom_all-v['kargo']-v['platform']-mal) if mal is not None else None
            satirlar.append({**v,'barkod':bk,'stopaj':stopaj,'kom_all':kom_all,'mal_birim':mb,
                             'maliyet':mal,'net':net,'marj':(net/v['ciro']) if (net is not None and v['ciro']) else None})
        satirlar.sort(key=lambda x:-x['ciro'])
        eksik=[x['ad'] or x['barkod'] for x in satirlar if x['maliyet'] is None]
        net_t=sum(x['net'] for x in satirlar if x['net'] is not None)
        ciro_t=sum(x['ciro'] for x in satirlar)
        return json.dumps(dict(gun=gun, satirlar=satirlar, eksik=eksik,
            desi=getattr(self,'_desi_bilgi',None),
            toplam=dict(ciro=ciro_t, kom=sum(x['kom'] for x in satirlar),
                        kom_t=sum(x['kom_t'] for x in satirlar),
                        kom_o=sum(x['kom_o'] for x in satirlar),
                        stopaj=sum(x['stopaj'] for x in satirlar),
                        kom_all=sum(x['kom_all'] for x in satirlar),
                        kargo=sum(x['kargo'] for x in satirlar), kargo_t=sum(x['kargo_t'] for x in satirlar),
                        platform=sum(x['platform'] for x in satirlar),
                        maliyet=sum(x['maliyet'] for x in satirlar if x['maliyet'] is not None),
                        net=net_t, marj=(net_t/ciro_t) if ciro_t else None,
                        adet=sum(x['adet'] for x in satirlar))), ensure_ascii=False)

    def urun_listesi(self, yenile=0):
        """Trendyol'daki TÜM satıştaki ürünler + maliyet/desi (Excel) + komisyon (tarife/hakediş)."""
        import requests, base64, time as _t
        if getattr(self,'_ulist',None) and not int(yenile):
            return self._ulist_json()
        cfg=json.load(open(ayar.TY_API))
        sid=cfg['seller_id']
        tok=base64.b64encode(f"{cfg['api_key']}:{cfg['api_secret']}".encode()).decode()
        H={'Authorization':f'Basic {tok}','User-Agent':f'{sid} - SelfIntegration'}
        A='https://apigw.trendyol.com/integration'
        # 1) fiyat/stok
        pr={}
        for page in range(30):
            r=requests.get(f'{A}/product/sellers/{sid}/products/approved/inventory-and-price',
                           params={'page':page,'size':100},headers=H,timeout=30).json()
            for p in r.get('content',[]):
                for v in p.get('variants',[]):
                    b=nb(v.get('barcode'))
                    if b: pr[b]=dict(fiyat=v.get('salePrice'), stok=v.get('quantity'), cid=p.get('contentId'))
            if page>=r.get('totalPages',1)-1: break
        # 2) başlık/kategori
        for page in range(30):
            r=requests.get(f'{A}/product/sellers/{sid}/products/approved',
                           params={'page':page,'size':100},headers=H,timeout=30).json()
            for p in r.get('content',[]):
                for v in (p.get('variants') or [{}]):
                    b=nb(v.get('barcode'))
                    if b and b in pr:
                        pr[b]['ad']=str(p.get('title') or '')[:70]
                        def _ad(x):
                            if isinstance(x,dict): return str(x.get('name') or '')
                            return str(x or '')
                        pr[b]['kategori']=_ad(p.get('category') or p.get('categoryName'))
                        pr[b]['marka']=_ad(p.get('brand'))
                        if v.get('commission') is not None: pr[b]['api_kom']=float(v['commission'])
            if page>=r.get('totalPages',1)-1: break
        # 3) komisyon: (a) tarife dosyası (güncel komisyon), (b) son hakedişte fiilen kesilen oran
        kom={}; kaynak={}
        for u in self.v.urunler:
            if u.get('guncel_kom'): kom[u['barkod']]=u['guncel_kom']; kaynak[u['barkod']]='tarife'
        simdi=_t.time()
        for g in (45,32,19,6):
            s_=int((simdi-g*86400)*1000); e_=int((simdi-max(0,g-13)*86400)*1000)
            page=0
            while True:
                r=requests.get(f'{A}/finance/che/sellers/{sid}/settlements',
                               params={'startDate':s_,'endDate':e_,'transactionType':'Sale','page':page,'size':500},
                               headers=H,timeout=30).json()
                for k in r.get('content',[]):
                    b=nb(k.get('barcode')); o=k.get('commissionRate')
                    if b and o and b not in kom: kom[b]=o; kaynak[b]='satış'
                if page>=r.get('totalPages',1)-1: break
                page+=1
        # kategori bazlı çıkarım: bilinen komisyonların kategori medyanı
        katkom={}
        for b,o in kom.items():
            k=(pr.get(b) or {}).get('kategori')
            if k: katkom.setdefault(k,[]).append(o)
        katmed={k:sorted(v)[len(v)//2] for k,v in katkom.items()}
        out=[]
        for b,d in pr.items():
            if b not in kom:
                km=katmed.get(d.get('kategori'))
                if km: kom[b]=km; kaynak[b]='kategori'
            g=self.v.girdi.get(b,{})
            km=g.get('kom_manuel'); kk='manuel'
            if km is None and d.get('api_kom') is not None: km=d['api_kom']; kk='api'
            if km is None: km=kom.get(b); kk=kaynak.get(b,'')
            out.append(dict(barkod=b, ad=d.get('ad') or '', kategori=d.get('kategori') or '',
                            marka=d.get('marka') or '', fiyat=d.get('fiyat'), stok=d.get('stok'),
                            kom=km, kom_kaynak=kk,
                            maliyet=g.get('maliyet'), desi=g.get('desi')))
        out.sort(key=lambda x:(x['maliyet'] is not None, -(x['stok'] or 0)))
        self._ulist=out
        self._meta_yaz(out)
        return self._ulist_json()

    def _meta_yaz(self, liste):
        """Ürün adı/kategori/API komisyonunu Maliyetler'e (B/M/N) işler. Tarife Takibi
        ekranı tarifesiz ürünleri buradan tanıyor — API'ye gitmeden."""
        try:
            wb=openpyxl.load_workbook(XLSX); ml=wb['Maliyetler']
            for c,bas in ((13,'Kategori'),(14,'Komisyon (API)'),(15,'Buybox Sırası'),(16,'Rakip Var')):
                if not ml.cell(1,c).value: ml.cell(1,c,bas)
            satir={}
            for r in range(2, ml.max_row+1):
                b=ml.cell(r,1).value
                if b: satir[nb(b)]=r
            n=0
            for x in liste:
                if not ((x.get('stok') or 0)>0): continue     # yalnız satıştaki ürünler
                r=satir.get(x['barkod'])
                if r is None:
                    r=ml.max_row+1; ml.cell(r,1,x['barkod']); satir[x['barkod']]=r
                if x.get('ad') and not ml.cell(r,2).value: ml.cell(r,2,x['ad'])
                if x.get('kategori'): ml.cell(r,13,x['kategori'])
                if x.get('kom') is not None: ml.cell(r,14,float(x['kom']))
                n+=1
            wb.save(XLSX); self.v.yukle()
            return n
        except Exception:
            return 0

    def _ulist_json(self):
        u=self._ulist
        return json.dumps(dict(urunler=u, ozet=dict(
            toplam=len(u), maliyetli=sum(1 for x in u if x['maliyet'] is not None),
            stokta=sum(1 for x in u if (x['stok'] or 0)>0),
            komsuz=sum(1 for x in u if not x['kom']),
            kom_manuel=sum(1 for x in u if x.get('kom_kaynak')=='manuel'))), ensure_ascii=False)

    # ================= HEPSİBURADA =================
    def _hb(self):
        import base64
        c=json.load(open(ayar.HB_API))
        tok=base64.b64encode(f"{c['merchant_id']}:{c['service_key']}".encode()).decode()
        return c['merchant_id'], {'Authorization':f'Basic {tok}','User-Agent':c['user_agent'],'Accept':'application/json'}

    # ───────────────────── FIRSAT PROGRAMLARI (flaş + yıldızlı ürün) ─────────────────────
    def _kom_at(self, bk, fiyat):
        """Barkodun VERİLEN FİYATTAKİ komisyon oranı. Önce barem tarifesi (fiyat düşünce
        komisyon da düşer — fırsat fiyatlarında bu çok önemli), yoksa ürün API'sinin oranı."""
        if fiyat is None: return None, ''
        u=self.v.bant.get(bk)
        if u:
            kom=u['kom1'] if self.v.donem==1 else u['kom2']
            I,J,K,L,M,N=u['limit']
            for lim,kk in ((I,kom[0]),(K,kom[1]),(M,kom[2])):
                if lim is not None and fiyat>=lim and kk is not None: return kk,'barem'
            return kom[3],'barem'
        k=self._api_komisyon().get(bk)
        return (k,'api') if k is not None else (None,'')

    def _kalem(self, bk, fiyat, etiket=''):
        """Bir fiyatın tam kesinti dökümü — ciro raporuyla birebir aynı kalemler."""
        g=self.v.girdi.get(bk,{}); mal=g.get('maliyet'); desi=g.get('desi') or 1
        kom,src=self._kom_at(bk, fiyat)
        if fiyat is None or kom is None or mal is None:
            return dict(etiket=etiket, fiyat=fiyat, kom=kom, eksik=True)
        komtl=fiyat*kom/100
        stopaj=fiyat*self.v.STOPAJ_ORAN
        kargo=self.v.kargo(fiyat, desi)
        net=fiyat-komtl-stopaj-kargo-PLATFORM-mal
        return dict(etiket=etiket, fiyat=fiyat, kom=kom, kom_kaynak=src, komtl=komtl,
                    stopaj=stopaj, kargo=kargo, platform=PLATFORM, maliyet=mal, desi=desi,
                    net=net, pct=(net/fiyat) if fiyat else None, eksik=False)

    def _son_tarife(self, tip):
        """Tarifeler/ arşivindeki EN YENİ dosyayı bulur (hafta klasörlerinin hepsine bakar)."""
        import glob
        d=sorted(glob.glob(os.path.join(TARIFE_ARSIV,'*',f'{tip}_*.xlsx')), key=os.path.getmtime)
        return d[-1] if d else None

    def firsat_listesi(self, yenile=0):
        """Flaş + yıldızlı ürün eşiklerini tek listede toplar, her eşiğin net kârını hesaplar."""
        if getattr(self,'_flist',None) and not int(yenile):
            return json.dumps(self._flist, ensure_ascii=False)
        def _f(x):
            if x is None: return None
            if isinstance(x,(int,float)): return float(x)
            try: return float(str(x).replace('.','').replace(',','.'))
            except Exception: return None
        urun={}
        def U(bk, ad='', kat='', stok=None, fiyat=None):
            x=urun.setdefault(bk, dict(barkod=bk, ad='', kategori='', stok=None, fiyat=None,
                                       yildiz=[None,None,None], flas24=None, flas3=None, slot=0, tarihler=[]))
            if ad and not x['ad']: x['ad']=ad
            if kat and not x['kategori']: x['kategori']=kat
            if stok is not None: x['stok']=stok
            if fiyat is not None and x['fiyat'] is None: x['fiyat']=fiyat
            return x
        kaynak={}
        y=self._son_tarife('yildiz')
        if y:
            kaynak['yildiz']=os.path.basename(y)
            ws=openpyxl.load_workbook(y, data_only=True)['YıldızlıÜrünEtiketleri']
            for r in range(2, ws.max_row+1):
                bk=str(ws.cell(r,2).value or '').strip()
                if not bk: continue
                x=U(bk, str(ws.cell(r,1).value or '')[:60], str(ws.cell(r,5).value or ''),
                    _f(ws.cell(r,7).value), _f(ws.cell(r,15).value))
                x['yildiz']=[_f(ws.cell(r,9).value), _f(ws.cell(r,11).value), _f(ws.cell(r,13).value)]
        fl=self._son_tarife('flas')
        if fl:
            kaynak['flas']=os.path.basename(fl)
            ws=openpyxl.load_workbook(fl, data_only=True)['TeklifÜrünleri']
            for r in range(2, ws.max_row+1):
                bk=str(ws.cell(r,2).value or '').strip()
                if not bk: continue
                x=U(bk, str(ws.cell(r,3).value or '')[:60], str(ws.cell(r,4).value or ''),
                    _f(ws.cell(r,6).value), _f(ws.cell(r,7).value))
                p24=_f(ws.cell(r,11).value); p3=_f(ws.cell(r,12).value)
                if p24 and (x['flas24'] is None or p24>x['flas24']): x['flas24']=p24
                if p3 and (x['flas3'] is None or p3>x['flas3']): x['flas3']=p3
                x['slot']+=1
                t=ws.cell(r,14).value
                if t and str(t) not in x['tarihler']: x['tarihler'].append(str(t))
        out=[]
        for bk,x in urun.items():
            g=self.v.girdi.get(bk,{})
            if x['fiyat'] is None: x['fiyat']=g.get('canli')
            sec=[self._kalem(bk, x['fiyat'], 'Mevcut fiyat')]
            for i,ad in enumerate(('1 ★','2 ★','3 ★')):
                if x['yildiz'][i]: sec.append(self._kalem(bk, x['yildiz'][i], ad))
            if x['flas24']: sec.append(self._kalem(bk, x['flas24'], 'Flaş 24 saat'))
            if x['flas3']:  sec.append(self._kalem(bk, x['flas3'],  'Flaş 3 saat'))
            gecerli=[s for s in sec[1:] if not s.get('eksik') and s.get('pct') is not None]
            eniyi=max(gecerli, key=lambda s:s['pct']) if gecerli else None
            out.append(dict(**{k:x[k] for k in ('barkod','ad','kategori','stok','fiyat','slot','tarihler')},
                            maliyet=g.get('maliyet'), desi=g.get('desi'),
                            secenekler=sec, eniyi=eniyi,
                            simdi_pct=sec[0].get('pct'), flas_var=bool(x['flas24'] or x['flas3']),
                            yildiz_var=any(x['yildiz'])))
        out.sort(key=lambda z:-(z['eniyi']['pct'] if z['eniyi'] else -9))
        self._flist=dict(urunler=out, kaynak=kaynak,
                         ozet=dict(toplam=len(out), flas=sum(1 for z in out if z['flas_var']),
                                   yildiz=sum(1 for z in out if z['yildiz_var']),
                                   maliyetsiz=sum(1 for z in out if z['maliyet'] is None)))
        return json.dumps(self._flist, ensure_ascii=False)

    def hb_urun_listesi(self, yenile=0):
        """HB ürün listesi = LISTING (satışta olanlar, 649) ∪ KATALOG (kendi ürünlerin, 868).
        Katalogda olmayan listing'lerin (HB kataloğuna satıcı olduğun ürünler) ad/barkodu
        sipariş defterinden tamamlanır. Maliyet/desi ortak Excel'den (barkod, yoksa SKU anahtarı)."""
        import requests
        if getattr(self,'_hblist',None) and not int(yenile): return self._hb_json()
        mid,H=self._hb()
        kat={}
        for page in range(40):
            r=requests.get(f'https://mpop.hepsiburada.com/product/api/products/all-products-of-merchant/{mid}',
                           params={'page':page,'size':100},headers=H,timeout=30).json()
            for d in r.get('data',[]):
                sku=str(d.get('hbSku') or '')
                if sku: kat[sku]=dict(barkod=str(d.get('barcode') or '').strip(),
                                      ad=str(d.get('productName') or '')[:70],
                                      msku=str(d.get('merchantSku') or ''),
                                      kategori=str(d.get('categoryName') or ''), marka=str(d.get('brand') or ''))
            if page>=(r.get('totalPages') or 1)-1: break
        fs={}
        for off in range(0,6000,100):
            r=requests.get(f'https://listing-external.hepsiburada.com/listings/merchantid/{mid}',
                           params={'limit':100,'offset':off},headers=H,timeout=30).json()
            L=r.get('listings',[])
            for l in L:
                fs[str(l.get('hepsiburadaSku'))]=dict(fiyat=l.get('price'), stok=l.get('availableStock'),
                                                      satista=bool(l.get('isSalable')), msku=str(l.get('merchantSku') or ''))
            if off+100>=(r.get('totalCount') or 0) or not L: break
        # KOMİSYON SERVİSİ: listing-external/commissions (50 SKU'luk gruplar; oran + vade + kampanya)
        kom={}
        skular=[k for k in fs]
        for i in range(0,len(skular),50):
            grup=skular[i:i+50]
            try:
                rr=requests.get(f'https://listing-external.hepsiburada.com/commissions/merchantid/{mid}',
                                params={'skuList':','.join(grup)},headers=H,timeout=30).json()
            except Exception: continue
            for x in (rr.get('commissions') or []):
                sk=str(x.get('hepsiburadaSku') or '')
                pr_=(x.get('pricings') or [])
                kmp=pr_[0] if pr_ else None
                kom[sk]=dict(oran=x.get('rate'), vade=x.get('paymentTermInDays'),
                             kamp_fiyat=(kmp or {}).get('finalPrice'),
                             kamp_bitis=str((kmp or {}).get('endDate') or '')[:10],
                             kamp_hb=sum((d.get('amount') or 0) for d in ((kmp or {}).get('debtors') or [])
                                         if str(d.get('debtor','')).lower().startswith('hepsi')))
        # sipariş defterinden ad/barkod tamamlama (katalogda olmayanlar için)
        try: defter=json.load(open(self.HBDEFTER))
        except Exception: defter={}
        sip={}
        for sp in defter.values():
            for k in sp.get('kalemler',[]):
                sk=str(k.get('sku') or '')
                if sk and sk not in sip: sip[sk]=dict(ad=k.get('ad') or '', barkod=nb(k.get('barkod')))
        out=[]
        for sku in set(list(kat)+list(fs)):
            d=kat.get(sku) or {}; f=fs.get(sku) or {}; sp=sip.get(sku) or {}
            barkod=nb(d.get('barkod') or sp.get('barkod') or '')
            ad=d.get('ad') or sp.get('ad') or ''
            anahtar=barkod or sku
            g=self.v.girdi.get(anahtar,{})
            km=kom.get(sku) or {}
            out.append(dict(barkod=barkod, anahtar=anahtar, ad=ad, sku=sku, msku=d.get('msku') or f.get('msku') or '',
                            kategori=d.get('kategori',''), marka=d.get('marka',''),
                            kom=km.get('oran'), vade=km.get('vade'), kamp_fiyat=km.get('kamp_fiyat'),
                            kamp_bitis=km.get('kamp_bitis'), kamp_hb=km.get('kamp_hb'),
                            fiyat=f.get('fiyat'), stok=f.get('stok'), satista=f.get('satista',False),
                            katalogda=sku in kat, satista_listing=sku in fs,
                            maliyet=g.get('maliyet'), desi=g.get('desi')))
        out.sort(key=lambda x:(x['maliyet'] is not None, -(x['stok'] or 0)))
        self._hblist=out
        return self._hb_json()

    def _hb_json(self):
        u=self._hblist
        return json.dumps(dict(urunler=u, ozet=dict(
            toplam=len(u), maliyetli=sum(1 for x in u if x['maliyet'] is not None),
            stokta=sum(1 for x in u if (x['stok'] or 0)>0),
            barkodsuz=sum(1 for x in u if not x['barkod']),
            komlu=sum(1 for x in u if x.get('kom')),
            kampanyali=sum(1 for x in u if x.get('kamp_fiyat')),
            listingde=sum(1 for x in u if x.get('satista_listing')),
            katalogsuz=sum(1 for x in u if x.get('satista_listing') and not x.get('katalogda')))), ensure_ascii=False)

    HBDEFTER=ayar.HB_DEFTER

    def _hb_defter_guncelle(self, log=None):
        """HB sipariş geçmişi: statü listelerinden (delivered/shipped/undelivered + aktif paketler)
        sipariş numaraları toplanır, her yeni sipariş için DETAY çekilir (kalem+komisyon+desi+barkod)
        ve yerel deftere yazılır. Detay bir kez çekilir, sonraki açılışlarda tekrar çekilmez."""
        import requests
        mid,H=self._hb()
        B=f'https://oms-external.hepsiburada.com'
        try: defter=json.load(open(self.HBDEFTER))
        except Exception: defter={}
        sipno=set()
        # 1) statü listeleri (sayfalı, sipariş no verir)
        for yol in ('delivered','shipped','undelivered'):
            off=0
            while True:
                try:
                    r=requests.get(f'{B}/packages/merchantid/{mid}/{yol}',params={'offset':off,'limit':50},headers=H,timeout=30).json()
                except Exception: break
                it=r.get('items') or []
                for x in it:
                    o=str(x.get('OrderNumber') or x.get('orderNumber') or '')
                    if o: sipno.add(o)
                off+=50
                if off>=(r.get('totalCount') or 0) or not it: break
        # 2) aktif paketler (henüz teslim olmamışlar)
        for st in ('Open','Packaged','Shipped','Delivered'):
            try:
                r=requests.get(f'{B}/packages/merchantid/{mid}',params={'status':st,'offset':0,'limit':100},headers=H,timeout=30).json()
            except Exception: continue
            if isinstance(r,list):
                for pk in r:
                    for o in (pk.get('orderNumbers') or []): sipno.add(str(o))
                    for it in (pk.get('items') or []):
                        o=str(it.get('orderNumber') or '')
                        if o: sipno.add(o)
        # 3) eksik siparişlerin detayını çek
        yeni=0
        for o in sipno:
            if o in defter: continue
            try:
                r=requests.get(f'{B}/orders/merchantid/{mid}/ordernumber/{o}',headers=H,timeout=30).json()
            except Exception: continue
            d=r[0] if isinstance(r,list) and r else (r if isinstance(r,dict) else None)
            if not d: continue
            kal=[]
            for k in (d.get('items') or []):
                kal.append(dict(
                    ad=str(k.get('name') or '')[:70],
                    barkod=str(k.get('productBarcode') or '').strip(),
                    sku=str(k.get('sku') or ''), msku=str(k.get('merchantSKU') or ''),
                    adet=k.get('quantity') or 1,
                    tutar=((k.get('totalPrice') or {}).get('amount')) or 0,
                    kom=((k.get('commission') or {}).get('amount')) or 0,
                    komor=k.get('commissionRate'),
                    desi=k.get('totalDeci') or k.get('deci') or 0,
                    bnpl=((k.get('bnplCommissionAmount') or {}).get('amount')) or 0,
                    iscilik=((k.get('unitLaborCost') or {}).get('amount')) or 0,
                    borclu=k.get('deptorDifferenceAmount') or 0,
                    hbisk=(((k.get('hbDiscount') or {}).get('totalPrice') or {}).get('amount')) or 0,
                    satisk=(((k.get('merchantDiscount') or {}).get('totalPrice') or {}).get('amount')) or 0,
                    kdvor=k.get('vatRate'),
                    durum=str(k.get('status') or '')))
            if kal:
                defter[o]=dict(tarih=str(d.get('orderDate') or (kal and '') or ''), kalemler=kal); yeni+=1
        try: json.dump(defter, open(self.HBDEFTER,'w'), ensure_ascii=False)
        except Exception: pass
        if log: log(f'HB defter: {len(defter)} sipariş (+{yeni} yeni)')
        return defter, yeni

    def hb_ciro(self, gun):
        """HB ürün bazlı ciro/komisyon/kargo/maliyet → net kâr (sipariş tarihi bazlı)."""
        import datetime as _dt
        gun=int(gun)
        if not getattr(self,'_hblist',None): self.hb_urun_listesi(0)
        bar2={x['barkod']:x for x in self._hblist if x['barkod']}
        bar2.update({x['anahtar']:x for x in self._hblist})
        sku2={x['sku']:x for x in self._hblist}
        defter,yeni=self._hb_defter_guncelle()
        sinir=_dt.datetime.now()-_dt.timedelta(days=gun)
        IPTAL=('Cancelled','CancelledByCustomer','CancelledByMerchant','CancelledBySap','Returned')
        urun={}
        def d(bk,ad=''):
            x=urun.setdefault(bk,{'ad':ad,'ciro':0.0,'kom':0.0,'kargo':0.0,'kargo_t':0.0,'adet':0,'indirim':0.0})
            if ad and not x['ad']: x['ad']=ad
            return x
        for onum,sp in defter.items():
            try: od=_dt.datetime.fromisoformat(str(sp.get('tarih'))[:19])
            except Exception: continue
            if od<sinir: continue
            gecerli=[k for k in sp.get('kalemler',[]) if k.get('durum') not in IPTAL]
            if not gecerli: continue
            # kargo ve hizmet bedeli SİPARİŞ başınadır → toplam desiden hesaplanıp ciro payıyla dağıtılır
            top_desi=sum((k.get('desi') or (self.v.girdi.get(nb(k.get('barkod')),{}).get('desi') or 1)) for k in gecerli)
            dd=int(max(1,min(20,round(top_desi or 1))))
            sip_kargo=self.v.hb_kargo.get(dd,106.8)+HB_HIZMET
            top_ciro=sum((k.get('tutar') or 0) for k in gecerli) or 1
            for k in gecerli:
                bk=nb(k.get('barkod')) if k.get('barkod') else (k.get('sku') or '?')
                u=bar2.get(bk) or sku2.get(k.get('sku'))
                x=d(bk, k.get('ad') or (u or {}).get('ad',''))
                tut=k.get('tutar') or 0
                x['ciro']+=tut
                # API komisyonu ADET BAŞINA ve KDV HARİÇ → ×adet ×1,20 (panel ekstresiyle doğrulandı)
                x['kom']+=(k.get('kom') or 0)*(k.get('adet') or 1)*HB_KOM_KDV
                x['indirim']+=k.get('hbisk') or 0             # HB'nin karşıladığı indirim → satıcıya iade
                x['adet']+=k.get('adet') or 0
                pay=sip_kargo*(tut/top_ciro)
                x['kargo']+=pay; x['kargo_t']+=pay
        satirlar=[]
        for bk,v in urun.items():
            stopaj=(v['ciro']/(1+KDV))*STOPAJ
            mb=self.v.girdi.get(bk,{}).get('maliyet')
            mal=(mb*v['adet']) if (mb is not None and v['adet']>0) else (0.0 if mb is not None else None)
            tahsilat=v['ciro']*HB_TAHSILAT
            kom_all=v['kom']+stopaj+tahsilat-v.get('indirim',0)   # HB indirimi kesintiyi azaltır
            net=(v['ciro']-kom_all-v['kargo']-mal) if mal is not None else None
            satirlar.append({**v,'barkod':bk,'stopaj':stopaj,'tahsilat':tahsilat,'kom_all':kom_all,'mal_birim':mb,'kom_t':0,
                             'maliyet':mal,'net':net,'marj':(net/v['ciro']) if (net is not None and v['ciro']) else None})
        satirlar.sort(key=lambda x:-x['ciro'])
        ciro_t=sum(x['ciro'] for x in satirlar); net_t=sum(x['net'] for x in satirlar if x['net'] is not None)
        return json.dumps(dict(gun=gun, satirlar=satirlar, defter=len(defter), yeni=yeni,
            eksik=[x['ad'] or x['barkod'] for x in satirlar if x['maliyet'] is None],
            toplam=dict(ciro=ciro_t, kom=sum(x['kom'] for x in satirlar), kom_t=0,
                        stopaj=sum(x['stopaj'] for x in satirlar), kom_all=sum(x['kom_all'] for x in satirlar),
                        tahsilat=sum(x['tahsilat'] for x in satirlar), indirim=sum(x.get('indirim',0) for x in satirlar),
                        kargo=sum(x['kargo'] for x in satirlar), kargo_t=sum(x['kargo_t'] for x in satirlar),
                        platform=sum(x['platform'] for x in satirlar),
                        maliyet=sum(x['maliyet'] for x in satirlar if x['maliyet'] is not None),
                        net=net_t, marj=(net_t/ciro_t) if ciro_t else None,
                        adet=sum(x['adet'] for x in satirlar))), ensure_ascii=False)

    def toplu_kaydet(self, kayitlar):
        """[{barkod,maliyet,desi},...] listesini TEK Excel açılışında yazar."""
        try:
            kay=json.loads(kayitlar) if isinstance(kayitlar,str) else kayitlar
        except Exception:
            kay=[]
        if not kay: return self._ulist_json()
        wb=openpyxl.load_workbook(XLSX); ml=wb['Maliyetler']
        satir={}
        for r in range(2, ml.max_row+1):
            b=ml.cell(r,1).value
            if b: satir[str(b).strip()]=r; satir[nb(b)]=r
        adlar={}
        for L in ((getattr(self,'_ulist',None) or []), (getattr(self,'_hblist',None) or [])):
            for x in L:
                k=x.get('anahtar') or x.get('barkod')
                if k and x.get('ad'): adlar[k]=x['ad']
        def _n(v):
            try:
                v=str(v).replace(',','.').strip()
                return float(v) if v not in ('','None','null') else None
            except Exception: return None
        n=0
        for k in kay:
            bk=nb(k.get('barkod'))
            if not bk: continue
            m=_n(k.get('maliyet')); d=_n(k.get('desi')); km=_n(k.get('kom'))
            if m is None and d is None and km is None: continue
            r=satir.get(bk)
            if r is None:
                r=ml.max_row+1; ml.cell(r,1,bk); ml.cell(r,2,adlar.get(bk,'')); satir[bk]=r
            if m is not None: ml.cell(r,3,m)
            if d is not None: ml.cell(r,4,d)
            if km is not None: ml.cell(r,10,km)
            ml.cell(r,9,'manuel giriş'); n+=1
            for L in ((getattr(self,'_ulist',None) or []), (getattr(self,'_hblist',None) or [])):
                for x in L:
                    if x.get('barkod')==bk or x.get('anahtar')==bk:
                        if m is not None: x['maliyet']=m
                        if d is not None: x['desi']=d
                        if km is not None: x['kom']=km; x['kom_kaynak']='manuel'
        wb.save(XLSX); self.v.yukle()
        out=json.loads(self._ulist_json()) if getattr(self,'_ulist',None) else {'ozet':{}}
        out['kaydedilen']=n
        if getattr(self,'_hblist',None): out['hb']=json.loads(self._hb_json())
        return json.dumps(out, ensure_ascii=False)

    def _fiyat_stok_tazele(self):
        """inventory-and-price → Maliyetler F(fiyat)/G(stok)/H(zaman). Buybox çekimiyle birlikte
        koşar; böylece Trendyol'da fiyat değiştiği anda panelde görünür (eskiden yalnız
        ⟳ Güncelle tazeliyordu, panel bayat fiyatla hesap yapıyordu)."""
        import requests, base64
        try:
            cfg=json.load(open(ayar.TY_API)); sid=cfg['seller_id']
            tok=base64.b64encode(f"{cfg['api_key']}:{cfg['api_secret']}".encode()).decode()
            H={'Authorization':f'Basic {tok}','User-Agent':f'{sid} - SelfIntegration'}
            pr={}
            for page in range(30):
                r=requests.get(f'https://apigw.trendyol.com/integration/product/sellers/{sid}/products/approved/inventory-and-price',
                               params={'page':page,'size':100},headers=H,timeout=30).json()
                for p in r.get('content',[]):
                    for v in p.get('variants',[]):
                        b=nb(v.get('barcode'))
                        if b: pr[b]=(v.get('salePrice'), v.get('quantity'))
                if page>=r.get('totalPages',1)-1: break
            if not pr: return 0
            wb=openpyxl.load_workbook(XLSX); ml=wb['Maliyetler']
            ts=f'{datetime.datetime.now():%Y-%m-%d %H:%M}'; n=0
            for r_ in range(2, ml.max_row+1):
                b=ml.cell(r_,1).value
                if not b: continue
                d=pr.get(nb(b))
                if not d: continue
                ml.cell(r_,6,d[0]); ml.cell(r_,7,d[1]); ml.cell(r_,8,ts); n+=1
            wb.save(XLSX); self.v.yukle()
            return n
        except Exception:
            return 0

    def buybox_cek(self, sadece_tarife=0):
        """Trendyol buybox servisi (10 barkodluk gruplar) → Maliyetler'e yazar:
        E = buybox fiyatı · L = 2. satıcı fiyatı · O = sıramız · P = rakip var mı.
        sadece_tarife=1 → STOKTAKİ ürünler (Tarife ekranı açılışı, ~8 istek),
        0 → satıştaki tüm ürünler (Ürünler ekranındaki düğme)."""
        import requests, base64
        tazelenen=self._fiyat_stok_tazele()      # önce canlı fiyat/stok, sonra buybox
        cfg=json.load(open(ayar.TY_API)); sid=cfg['seller_id']
        tok=base64.b64encode(f"{cfg['api_key']}:{cfg['api_secret']}".encode()).decode()
        H={'Authorization':f'Basic {tok}','User-Agent':f'{sid} - SelfIntegration','Content-Type':'application/json'}
        if sadece_tarife:
            barkodlar=[u['barkod'] for u in self.v.hesapla() if u.get('barkod')]
        else:
            if not getattr(self,'_ulist',None): self.urun_listesi(0)
            barkodlar=[x['barkod'] for x in self._ulist if x.get('barkod')]
        sonuc={}
        for i in range(0,len(barkodlar),10):
            grup=barkodlar[i:i+10]
            try:
                r=requests.post(f'https://apigw.trendyol.com/integration/product/sellers/{sid}/products/buybox-information',
                                json={'barcodes':grup},headers=H,timeout=30)
                if r.status_code!=200: continue
                for x in r.json().get('buyboxInfo',[]):
                    b=nb(x.get('barcode'))
                    # tek satıcı olduğumuz ürünlerde buyboxPrice boş gelebiliyor; o kaydı da
                    # tutuyoruz, yoksa "tek satıcı" durumu hiç görünmez
                    if b: sonuc[b]=dict(fiyat=x.get('buyboxPrice'), sira=x.get('buyboxOrder'),
                                        coklu=bool(x.get('hasMultipleSeller')),
                                        ikinci=x.get('secondBuyboxPrice'))
            except Exception: continue
        wb=openpyxl.load_workbook(XLSX); ml=wb['Maliyetler']
        satir={}
        for r_ in range(2, ml.max_row+1):
            b=ml.cell(r_,1).value
            if b: satir[nb(b)]=r_
        adlar={x['barkod']:x['ad'] for x in (getattr(self,'_ulist',None) or []) if x.get('barkod')}
        n=0
        for c,bas in ((13,'Kategori'),(14,'Komisyon (API)'),(15,'Buybox Sırası'),(16,'Rakip Var')):
            if not ml.cell(1,c).value: ml.cell(1,c,bas)
        for b,d in sonuc.items():
            r_=satir.get(b)
            if r_ is None:
                r_=ml.max_row+1; ml.cell(r_,1,b); ml.cell(r_,2,adlar.get(b,'')); satir[b]=r_
            if d.get('fiyat') is not None: ml.cell(r_,5,d['fiyat'])
            # ikinci fiyat/sıra/rakip HER SEFERİNDE yazılır (boşsa temizlenir) — eski değer
            # kalırsa "fiyat düşük kalmış" uyarısı hayalet veriyle çalışır
            ml.cell(r_,12,d.get('ikinci'))
            ml.cell(r_,15,d.get('sira'))
            ml.cell(r_,16, 1 if d.get('coklu') else 0)
            n+=1
        wb.save(XLSX); self.v.yukle()
        for x in (getattr(self,'_ulist',None) or []):
            d=sonuc.get(x.get('barkod'))
            if d: x['buybox']=d['fiyat']; x['bb_sira']=d['sira']; x['bb_coklu']=d['coklu']
        return json.dumps(dict(cekilen=len(sonuc), yazilan=n, tazelenen=tazelenen,
                               durum=json.loads(self.durum())), ensure_ascii=False)

    def urun_kaydet(self, barkod, maliyet, desi):
        """Ürünler ekranından maliyet/desi girişi → Excel + hafızadaki listeye işlenir."""
        wb=openpyxl.load_workbook(XLSX); ml=wb['Maliyetler']; satir=None; barkod=nb(barkod)
        for r in range(2, ml.max_row+1):
            if nb(ml.cell(r,1).value)==barkod: satir=r; break
        if satir is None:
            satir=ml.max_row+1; ml.cell(satir,1,str(barkod))
            for x in (getattr(self,'_ulist',None) or []):
                if x['barkod']==str(barkod): ml.cell(satir,2,x['ad']); break
        def _n(v):
            try: return float(str(v).replace(',','.')) if str(v).strip() not in ('','None') else None
            except Exception: return None
        m=_n(maliyet); d=_n(desi)
        if m is not None: ml.cell(satir,3,m)
        if d is not None: ml.cell(satir,4,d)
        ml.cell(satir,9,'manuel giriş')
        wb.save(XLSX); self.v.yukle()
        for x in (getattr(self,'_ulist',None) or []):
            if x['barkod']==str(barkod):
                if m is not None: x['maliyet']=m
                if d is not None: x['desi']=d
        return self._ulist_json()

    def maliyet_kaydet(self, barkod, maliyet):
        """Ciro ekranından hızlı maliyet girişi → Excel'e yazar, kârlılığı API'siz yeniden hesaplar."""
        wb=openpyxl.load_workbook(XLSX); ml=wb['Maliyetler']; satir=None; barkod=nb(barkod)
        for r in range(2, ml.max_row+1):
            if nb(ml.cell(r,1).value)==barkod: satir=r; break
        if satir is None:
            satir=ml.max_row+1; ml.cell(satir,1,str(barkod))
        try: ml.cell(satir,3, float(str(maliyet).replace(',','.')))
        except Exception: return json.dumps({'ok':False},ensure_ascii=False)
        ml.cell(satir,9,'manuel giriş')
        wb.save(XLSX); self.v.yukle()
        return self._ciro_sonuc()

    def fatura_maliyet(self):
        """Downloads'taki e-fatura PDF'lerinden maliyetleri çekip Excel'e yazar."""
        import importlib.util
        satir=[]
        try:
            spec=importlib.util.spec_from_file_location('fm',ayar.FATURA_PY)
            fm=importlib.util.module_from_spec(spec); spec.loader.exec_module(fm)
            r=fm.calistir(log=lambda m: satir.append(str(m)))
        except Exception as e:
            r={'yazilan':0}; satir.append(f'HATA: {e}')
        self.v.yukle()
        return json.dumps(dict(log='\n'.join(satir)[-3000:], sonuc=r, durum=json.loads(self.durum())), ensure_ascii=False)

    def guncelle(self):
        try:
            if getattr(ayar,'DONMUS',False):
                import runpy, io, contextlib
                _y=io.StringIO(); _a=sys.argv
                sys.argv=[ayar.GUNCELLE_PY,'--no-open']
                try:
                    with contextlib.redirect_stdout(_y), contextlib.redirect_stderr(_y):
                        runpy.run_path(ayar.GUNCELLE_PY, run_name='__main__')
                finally:
                    sys.argv=_a
                log=_y.getvalue()
            else:
                r=subprocess.run([sys.executable,ayar.GUNCELLE_PY,'--no-open'],
                                 capture_output=True,text=True,timeout=600)
                log=(r.stdout or '')+(r.stderr or '')
        except Exception as e:
            log=f'HATA: {e}'
        self.v.yukle()
        # Günlük ekranda GÖSTERİLMİYOR (kullanıcı isteği 2026-09-01) — dosyaya yazılır.
        try:
            with open(os.path.join(os.path.dirname(XLSX),'_guncelleme.log'),'a',encoding='utf-8') as f:
                f.write(f"\n===== {datetime.datetime.now():%Y-%m-%d %H:%M:%S} =====\n{log}")
        except Exception: pass
        return json.dumps(dict(durum=json.loads(self.durum())), ensure_ascii=False)

HTML='''<!DOCTYPE html><html lang="tr"><head><meta charset="utf-8"><style>
:root{--bg:#f4f5f9;--card:#ffffff;--line:#e7e9f0;--ink:#101828;--mut:#697386;
--or:#f27a1a;--or2:#ff9950;--yes:#059669;--no:#dc2626;--warn:#d97706;--blue:#2563eb;
--sh:0 1px 3px rgba(16,24,40,.07),0 10px 28px rgba(16,24,40,.06)}
*{box-sizing:border-box;margin:0}
body{background:var(--bg);color:var(--ink);font:13px/1.5 -apple-system,'SF Pro Text',Helvetica,sans-serif;
height:100vh;overflow:hidden}
.app{display:flex;height:100vh}
.ana{flex:1;display:flex;flex-direction:column;min-width:0}
.rail{width:212px;background:#fff;border-right:1px solid var(--line);display:flex;flex-direction:column;padding:14px 10px;gap:4px;overflow:auto}
.rail .marka{display:flex;align-items:center;gap:9px;padding:4px 6px 14px}
.rail .grp{font-size:10px;color:var(--mut);text-transform:uppercase;letter-spacing:1px;font-weight:800;padding:14px 10px 6px;display:flex;align-items:center;gap:7px}
.rail .grp i{width:7px;height:7px;border-radius:99px;display:inline-block}
.rail a{display:flex;align-items:center;gap:9px;padding:9px 11px;border-radius:10px;color:var(--ink);
font-size:12.5px;font-weight:650;cursor:pointer;transition:.13s;text-decoration:none}
.rail a:hover{background:#f4f5f9}
.rail a.on{background:linear-gradient(135deg,var(--or),#e8650a);color:#fff;box-shadow:0 3px 10px #f27a1a4a}
.rail a.pas{opacity:.45;cursor:default}
.rail a.pas:hover{background:none}
.rail .alt{margin-top:auto;font-size:10px;color:var(--mut);padding:10px}
*{font-variant-numeric:tabular-nums}
::-webkit-scrollbar{width:9px}::-webkit-scrollbar-thumb{background:#d4d9e4;border-radius:99px}
header{display:flex;align-items:center;gap:16px;padding:13px 22px;background:#fff;border-bottom:1px solid var(--line)}
.logo{width:34px;height:34px;border-radius:10px;background:linear-gradient(135deg,var(--or),#ff6d00);
display:grid;place-items:center;font-weight:900;color:#fff;font-size:17px;box-shadow:0 4px 12px #f27a1a4d}
header h1{font-size:15px;font-weight:800}header .sub{font-size:11px;color:var(--mut)}
.sp{flex:1}
.btn{border:0;border-radius:10px;padding:9px 16px;font-weight:700;font-size:12px;cursor:pointer;
background:linear-gradient(135deg,var(--or),#e8650a);color:#fff;box-shadow:0 4px 12px #f27a1a40;transition:.15s}
.btn:hover{transform:translateY(-1px)}.btn:disabled{opacity:.5}
.btn.ghost{background:#fff;color:var(--ink);box-shadow:none;border:1px solid var(--line)}
.seg{display:flex;background:#eceef4;border-radius:10px;padding:3px;gap:2px}
.seg div{padding:6px 14px;border-radius:8px;cursor:pointer;color:var(--mut);font-size:12px;font-weight:650;transition:.15s}
.seg div.on{background:#fff;color:var(--or);font-weight:800;box-shadow:0 1px 4px rgba(16,24,40,.12)}
main{flex:1;display:flex;min-height:0}
#sol{width:42%;min-width:410px;display:flex;flex-direction:column;padding:14px 6px 0 14px}
#ara{margin:0 8px 12px 0;padding:10px 14px;background:#fff;border:1px solid var(--line);border-radius:12px;width:calc(100% - 8px);font-size:13px;outline:none}
#ara:focus{border-color:var(--or)}
#liste{flex:1;overflow:auto;padding:2px 8px 14px 2px}
.urow{position:relative;background:var(--card);border:1px solid var(--line);border-radius:14px;padding:12px 14px 12px 18px;
margin-bottom:9px;cursor:pointer;transition:.15s;display:grid;grid-template-columns:1fr auto;gap:2px 12px;box-shadow:var(--sh)}
.urow:before{content:'';position:absolute;left:0;top:10px;bottom:10px;width:4px;border-radius:99px;background:var(--kt,#cbd5e1)}
.urow:hover{transform:translateY(-1px);box-shadow:0 4px 8px rgba(16,24,40,.09),0 14px 34px rgba(16,24,40,.08)}
.urow.on{border-color:var(--or);box-shadow:0 0 0 2px #f27a1a33,var(--sh)}
.urow .ad{font-weight:750;font-size:12.5px}
.urow .alt-s{color:var(--mut);font-size:11px;margin-top:2px}
.urow .fy{text-align:right;font-weight:850;font-size:15px}
.urow .kp{text-align:right;font-size:11px;font-weight:750}
.badge{display:inline-flex;align-items:center;gap:5px;font-size:10px;padding:3px 9px;border-radius:99px;font-weight:750;margin-top:6px}
.badge:before{content:'';width:6px;height:6px;border-radius:99px;background:currentColor}
.bg-g{background:#ecfdf3;color:#047857}.bg-r{background:#fef3f2;color:#b42318}.bg-w{background:#fffaeb;color:#b54708}
.bg-n{background:#f1f5f9;color:#475569}
.grpb{font-size:10px;font-weight:850;color:var(--mut);text-transform:uppercase;letter-spacing:.05em;padding:14px 4px 6px;border-top:1px solid var(--line);margin-top:8px}
.grpb:first-child{border-top:0;margin-top:0;padding-top:4px}
.tz{background:#f1f5f9;color:#475569;border-radius:5px;padding:1px 5px;font-size:10px;font-weight:700}
.g{color:var(--yes)}.r{color:var(--no)}.w{color:var(--warn)}.mut{color:var(--mut)}
#sag{flex:1;overflow:auto;padding:16px 20px}
.card{background:var(--card);border:1px solid var(--line);border-radius:16px;padding:18px;margin-bottom:14px;box-shadow:var(--sh)}
.card h2{font-size:11px;margin-bottom:12px;color:var(--mut);text-transform:uppercase;letter-spacing:1.1px;font-weight:750}
.kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:10px}
.kpi{border-radius:14px;padding:13px 15px;background:#f8fafc}
.kpi.k-g{background:#ecfdf3}.kpi.k-r{background:#fef3f2}.kpi.k-w{background:#fffaeb}.kpi.k-p{background:#fff4ed}
.kpi .t{font-size:10px;color:var(--mut);text-transform:uppercase;letter-spacing:.6px;font-weight:700}
.kpi .v{font-size:19px;font-weight:850;margin:4px 0 2px}
.kpi.k-g .v{color:#047857}.kpi.k-r .v{color:#b42318}.kpi.k-w .v{color:#b54708}.kpi.k-p .v{color:var(--or)}
.kpi .d{font-size:10.5px;color:var(--mut)}
svg text{font-family:-apple-system,'SF Pro Text',sans-serif}
input[type=number]{width:120px;background:#fff;border:1px solid var(--line);border-radius:10px;padding:9px 11px;font-size:13px;outline:none}
input[type=number]:focus{border-color:var(--or)}
label{font-size:10.5px;color:var(--mut);display:block;margin-bottom:5px;text-transform:uppercase;letter-spacing:.5px;font-weight:700}
.row2{display:flex;gap:16px;flex-wrap:wrap;align-items:flex-end}
input[type=range]{width:100%;-webkit-appearance:none;height:6px;border-radius:99px;background:#e6e9f2;outline:none}
input[type=range]::-webkit-slider-thumb{-webkit-appearance:none;width:22px;height:22px;border-radius:99px;
background:linear-gradient(135deg,var(--or),#ff6d00);cursor:grab;box-shadow:0 3px 10px #f27a1a66;border:3px solid #fff}
#simval{font-size:15px;font-weight:850}
#log{white-space:pre-wrap;font:10px 'SF Mono',Menlo,monospace;color:var(--mut);max-height:170px;overflow:auto;background:#f8fafc;border-radius:10px;padding:10px}
.bos{display:grid;place-items:center;height:70%;color:var(--mut)}
.bos .big{font-size:44px;margin-bottom:10px;opacity:.5}
table{width:100%;border-collapse:collapse;font-size:12px}
th{color:var(--mut);text-align:left;padding:9px 6px;font-size:11px;text-transform:uppercase;letter-spacing:.4px}
td{padding:9px 6px;border-top:1px solid #f0f2f7}
</style></head><body>
<div class="app">
<nav class="rail">
  <div class="marka"><div class="logo">P</div><div><div style="font-size:13px;font-weight:800">Pazaryeri</div>
  <div style="font-size:10px;color:var(--mut)">kârlılık paneli</div></div></div>
  <div class="grp"><i style="background:#f27a1a"></i>Trendyol</div>
  <a data-g="karlilik" class="on">🎯 Buybox & Fiyat</a>
  <a data-g="ciro">📈 Ciro Raporu</a>
  <a data-g="urunler">🗂 Ürünler</a>
  <a data-g="firsat">⚡ Fırsat Programları</a>
  <div class="grp"><i style="background:#ff6000"></i>Hepsiburada</div>
  <a data-g="hb_ciro">📈 Ciro Raporu</a>
  <a data-g="hb_urunler">🗂 Ürünler</a>
  <div class="alt" id="railalt"></div>
</nav>
<div class="ana">
<header><div id="baslik" style="font-size:15px;font-weight:800">Buybox &amp; Fiyat Kontrol</div>
<div class="sub" id="tarih"></div>
<div class="seg" id="donem"><div data-d="1">Dönem 1</div><div data-d="2">Dönem 2</div></div>
<span class="sp"></span>
<button class="btn ghost" onclick="guncelle()" id="gbtn">⟳ Güncelle</button></header>
<main><div id="sol">
<input id="ara" placeholder="🔍  Ürün, barkod, kategori veya durum ara (ör. fiyat düşük, tek satıcı)" oninput="ciz()">
<div id="bbdurum" class="mut" style="font-size:11px;padding:5px 3px 2px"></div>
<div id="liste"></div></div>
<div id="sag"><div class="bos"><div><div class="big">🛒</div>Soldan bir ürün seç</div></div></div>
<div id="ciro" style="display:none;flex:1;overflow:auto;padding:18px 20px"></div>
<div id="urunler" style="display:none;flex:1;overflow:auto;padding:18px 20px"></div>
<div id="firsat" style="display:none;flex:1;overflow:auto;padding:18px 20px"></div>
<div id="hb" style="display:none;flex:1;overflow:auto;padding:18px 20px"></div></main></div></div>
<script>
const PLATFORM=13.19, STOPAJ_ORAN=0.01/1.20, TOL=0.01;   // Trendyol platform hizmet bedeli, gönderi başına (KDV dahil)
let S=null, secili=null;
const tl=v=>v==null?'—':v.toLocaleString('tr-TR',{maximumFractionDigits:2})+' ₺';
const tk=v=>v==null?'—':(Math.abs(v)>=1000?(v/1000).toLocaleString('tr-TR',{maximumFractionDigits:1})+'K':v.toLocaleString('tr-TR',{maximumFractionDigits:0}));
const pc=v=>v==null?'—':(v*100).toLocaleString('tr-TR',{maximumFractionDigits:1})+'%';
const cls=(p,min)=>p==null?'mut':(p>=min?'g':(p>=0?'w':'r'));
let GIZLI=true, cAra='';    // sunum için varsayılan GİZLİ — göz düğmesiyle açılır
function cAraGir(el){const p=el.selectionStart;cAra=el.value.toLocaleLowerCase('tr');ciroCiz();
  const y=document.getElementById('cara');if(y){y.focus();y.setSelectionRange(p,p);}}
const tlm=v=>GIZLI?'***':tl(v);            // ciro ekranında tutar maskesi
function gizliDegis(){GIZLI=!GIZLI;ciroCiz();}
const KATR={};const PALET=['#f27a1a','#2563eb','#059669','#9333ea','#0891b2','#db2777','#ca8a04'];
function katRenk(k){if(!KATR[k])KATR[k]=PALET[Object.keys(KATR).length%PALET.length];return KATR[k];}
async function yukle(){S=JSON.parse(await window.pywebview.api.durum());ciz();}
// Liste artık BUYBOX DURUMUNA göre gruplu: önce aksiyon gerekenler.
const BBG={1:'💸 Kâr kabul aralığının altında — fiyat yükselt',2:'⚔️ Buybox rakipte — aksiyon var',
           3:'⚠️ Aksiyon yok / veri eksik',4:'🏆 Buybox bizde — kâr kabul aralığında',5:'👤 Tek satıcı'};
function bbRozetSinif(u){
  return {dusuk:'bg-r',sikisik:'bg-w',bizde:'bg-g',ustunde:'bg-g',maliyetsiz:'bg-w',tek:'bg-n',veriyok:'bg-n',rakipte_ucuz:'bg-w'}[u.bb_durum]
    || (u.durum.startsWith('✓')?'bg-g':(u.durum.startsWith('BUYBOX ALTI')?'bg-r':'bg-w'));
}
function bbSatir(u){
  const bant=`kabul %${(u.minp*100).toFixed(0)}–%${(u.tavan*100).toFixed(0)}`;
  if(u.bb_durum==='tek') return 'tek satıcıyız · rakip yok'+(u.canli_pct!=null?` · kâr ${pc(u.canli_pct)}`:'');
  if(u.bb_durum==='veriyok') return 'buybox verisi yok';
  if(u.bb_durum==='maliyetsiz') return 'buybox bizde · maliyet girilmemiş, kâr ölçülemiyor';
  if(u.bb_durum==='dusuk') return `kâr <b class="r">${pc(u.canli_pct)}</b> · ${bant} → ${tl(u.hedef)}`
    +(u.kisit==='rakip'?` <span class="tz">2. satıcı sınırladı</span>`:'');
  if(u.bb_durum==='sikisik') return `kâr <b class="r">${pc(u.canli_pct)}</b> · ${bant} ama üst fiyat ${tl(u.buybox2)} yer bırakmıyor`;
  if(u.bb_durum==='ustunde') return `kâr <b>${pc(u.canli_pct)}</b> · tavan %${(u.tavan*100).toFixed(0)} üstünde`;
  if(u.bb_durum==='rakipte_ucuz') return `${u.bb_sira}. sıradayız ama fiyatımız (${tl(u.canli)}) buybox'tan (${tl(u.buybox)}) DÜŞÜK`;
  if(u.bb_durum==='bizde') return `kâr ${pc(u.canli_pct)} · ${bant} ✓${u.buybox2?' · üst fiyat '+tl(u.buybox2):''}`;
  return `${u.bb_sira?u.bb_sira+'. sıradayız':'buybox rakipte'} · buybox ${tl(u.buybox)}`;
}
function ciz(){
  document.querySelectorAll('#donem div').forEach(e=>e.classList.toggle('on',+e.dataset.d===S.donem));
  document.getElementById('tarih').textContent=S.tarih[S.donem-1]||'';
  const q=(document.getElementById('ara').value||'').toLocaleLowerCase('tr');
  let h='',gp=null;
  for(const u of S.urunler){
    // arama ÖNERİ/BUYBOX DURUMUNU da kapsıyor: "fiyat düşük", "tek satıcı",
    // "buybox altına gir", "barem düşür", "maliyet girilmemiş"
    if(q && !(u.ad+u.barkod+u.kategori+' '+u.durum).toLocaleLowerCase('tr').includes(q)) continue;
    if(u.grup!==gp){gp=u.grup;h+=`<div class="grpb">${BBG[gp]||''}</div>`;}
    let fy,kp,kpc;
    if(u.grup===1){fy=tl(u.hedef);kp=(u.kacan!=null?'+'+tl(u.kacan)+' kâr/adet':'yükselt');kpc='g';}
    else if(u.grup===2||u.grup===3){fy=tl(u.oneri);kp=pc(u.oneri_pct)+' kâr';kpc=cls(u.oneri_pct,u.minp);}
    else{fy=tl(u.canli);kp='canlı fiyat';kpc='mut';}
    h+=`<div class="urow ${secili===u.barkod?'on':''}" style="--kt:${katRenk(u.kategori)}" onclick="sec('${u.barkod}')">
    <div><div class="ad">${(u.ad||u.barkod).slice(0,40)}</div>
    <div class="alt-s">${u.kategori||'—'} · maliyet ${tl(u.maliyet)} · canlı ${tl(u.canli)} / ${u.stok??'—'} stok
    ${u.tarifesiz?'· <span class="tz">tarife yok</span>':''}</div>
    <div class="alt-s">${bbSatir(u)}</div>
    <span class="badge ${bbRozetSinif(u)}">${u.durum.split('→')[0].trim()}</span></div>
    <div><div class="fy">${fy}</div>
    <div class="kp ${kpc}">${kp}</div></div></div>`;}
  document.getElementById('liste').innerHTML=h||'<div class="mut" style="padding:30px;text-align:center">Sonuç yok</div>';
  if(secili) detay(S.urunler.find(x=>x.barkod===secili));
}
function sec(bk){secili=bk;ciz();}
function komRenk(kom,koms){
  const mn=Math.min(...koms),mx=Math.max(...koms);
  const t=mx>mn?(kom-mn)/(mx-mn):0;   // 0=düşük kom (iyi) 1=yüksek
  if(t<0.34)return['#d1fae5','#047857'];if(t<0.67)return['#fef3c7','#b45309'];return['#fee2e2','#b91c1c'];
}
function fiyatEkseni(u){
  const bs=u.bantlar, koms=bs.map(b=>b.kom??0);
  const lo=Math.max(1,(bs[3].ust||100)*0.72);
  const hi=Math.max(bs[0].fiyat||0,u.buybox||0,u.buybox2||0,u.plus?u.plus.limit:0)*1.06;
  const X=f=>40+(Math.min(Math.max(f,lo),hi)-lo)/(hi-lo)*920;
  let sv=`<svg viewBox="0 0 1000 128" style="width:100%">`;
  const seg=[[lo,bs[3].ust,bs[3].kom],[bs[2].alt,bs[2].ust,bs[2].kom],[bs[1].alt,bs[1].ust,bs[1].kom],[bs[0].alt,hi,bs[0].kom]];
  for(const [a,b,k] of seg){
    if(a==null||b==null||k==null)continue;
    const[f,t]=komRenk(k,koms);
    sv+=`<rect x="${X(a)}" y="56" width="${Math.max(2,X(b)-X(a)-2)}" height="30" rx="7" fill="${f}"/>`;
    if(X(b)-X(a)>52)sv+=`<text x="${(X(a)+X(b))/2}" y="76" text-anchor="middle" font-size="13" font-weight="800" fill="${t}">%${k}</text>`;}
  sv+=`<text x="40" y="122" font-size="10" fill="#697386">${tk(lo)} ₺</text><text x="960" y="122" text-anchor="end" font-size="10" fill="#697386">${tk(hi)} ₺</text>`;
  if(u.canli){sv+=`<line x1="${X(u.canli)}" y1="30" x2="${X(u.canli)}" y2="56" stroke="#2563eb" stroke-width="2"/>
   <circle cx="${X(u.canli)}" cy="26" r="5" fill="#2563eb"/>
   <text x="${X(u.canli)}" y="14" text-anchor="middle" font-size="11" font-weight="750" fill="#2563eb">Canlı ${tk(u.canli)}</text>`;}
  if(u.buybox){sv+=`<line x1="${X(u.buybox)}" y1="34" x2="${X(u.buybox)}" y2="92" stroke="#dc2626" stroke-width="2" stroke-dasharray="4 3"/>
   <text x="${X(u.buybox)}" y="30" text-anchor="middle" font-size="11" font-weight="750" fill="#dc2626">Buybox ${tk(u.buybox)}</text>`;}
  if(u.buybox2){sv+=`<line x1="${X(u.buybox2)}" y1="42" x2="${X(u.buybox2)}" y2="92" stroke="#f472b6" stroke-width="1.5" stroke-dasharray="2 3"/>
   <text x="${X(u.buybox2)}" y="40" text-anchor="middle" font-size="10" font-weight="700" fill="#db2777">2. ${tk(u.buybox2)}</text>`;}
  if(u.oneri){sv+=`<path d="M ${X(u.oneri)} 86 l -7 14 h 14 z" fill="#059669"/>
   <text x="${X(u.oneri)}" y="114" text-anchor="middle" font-size="11" font-weight="800" fill="#059669">Öneri ${tk(u.oneri)}</text>`;}
  if(u.plus&&u.plus.limit){sv+=`<rect x="${X(u.plus.limit)-5}" y="88" width="10" height="10" rx="2" transform="rotate(45 ${X(u.plus.limit)} 93)" fill="#f27a1a"/>
   <text x="${X(u.plus.limit)}" y="114" text-anchor="middle" font-size="10" font-weight="750" fill="#f27a1a">Plus</text>`;}
  // imleç takibi: barem üzerinde gezerken komisyon + kâr canlı görünsün
  sv+=`<g id="fecross" style="display:none;pointer-events:none">
   <line id="feline" y1="20" y2="100" stroke="#101828" stroke-width="1" stroke-dasharray="3 3" opacity=".5"/></g>
   <rect x="40" y="20" width="920" height="80" fill="transparent" style="cursor:crosshair"
    onmousemove="feHover(event)" onmouseleave="feHoverOff()"/>`;
  window._fe={lo,hi,L:40,W:1000,R:40};
  return `<div id="fewrap" style="position:relative">${sv}</svg>
   <div id="fetip" style="position:absolute;display:none;pointer-events:none;background:#101828;color:#fff;
    padding:7px 11px;border-radius:9px;font-size:11px;line-height:1.7;white-space:nowrap;
    box-shadow:0 8px 24px rgba(16,24,40,.3);z-index:5"></div></div>`;
}
function feHover(e){
  const u=window._u, g=window._fe; if(!u||!g) return;
  const svg=document.querySelector('#fewrap svg'); if(!svg) return;
  const r=svg.getBoundingClientRect(), sc=r.width/g.W;
  let xv=Math.min(Math.max((e.clientX-r.left)/sc, g.L), g.W-g.R);
  const f=Math.round(g.lo+(xv-g.L)/(g.W-g.L-g.R)*(g.hi-g.lo));
  const k=kbKom(u,f), kar=kbKar(u,f), p=kbPct(u,f);
  if(k==null){feHoverOff();return;}
  const c=document.getElementById('fecross'); c.style.display='';
  const ln=document.getElementById('feline'); ln.setAttribute('x1',xv); ln.setAttribute('x2',xv);
  const renk=p==null?'#fff':(p>=u.minp?'#34d399':(p>=0?'#fbbf24':'#f87171'));
  const t=document.getElementById('fetip'); t.style.display='block';
  t.innerHTML=`<b style="font-size:12px">${tl(f)}</b><br>komisyon <b>%${k}</b>`
   +(kar==null?'<br><span style="opacity:.65">maliyet gir</span>'
     :`<br>kâr <b style="color:${renk}">${tl(kar)}</b> · <b style="color:${renk}">${pc(p)}</b>`)
   +(u.buybox?`<br><span style="opacity:.65">buybox ${f<=u.buybox-1?'ALINIR':'alınmaz'}</span>`:'');
  const tw=t.offsetWidth||150;
  t.style.left=Math.max(0,Math.min(r.width-tw, xv*sc+14))+'px';
  t.style.top=(r.height*0.62)+'px';
}
function feHoverOff(){
  const c=document.getElementById('fecross'), t=document.getElementById('fetip');
  if(c)c.style.display='none'; if(t)t.style.display='none';
}
function kbKom(u,f){if(u.tarifesiz)return u.kom_tek;
  for(const b of u.bantlar){if((b.alt==null||f>=b.alt)&&(b.ust==null||f<=b.ust))return b.kom;}return null;}
function kbKar(u,f){const k=kbKom(u,f);if(k==null||u.maliyet==null)return null;
  const kargo=f<=199.99?47.4:(f<=399.99?85.4:u.kargo_desi_tl);
  return f-f*k/100-f*STOPAJ_ORAN-kargo-PLATFORM-u.maliyet;}
function kbPct(u,f){const kar=kbKar(u,f);return kar==null?null:kar/f;}
function karBasamak(u){
  // Kâr% fiyata göre: bir bandın içinde artar, bant sınırında komisyon zıpladığı için DÜŞER.
  // Testere dişi bu yüzden — asıl karar bilgisi o sıçramanın nerede olduğu.
  if(u.maliyet==null) return '<div class="mut" style="padding:20px;text-align:center">Maliyet girilmeden kâr eğrisi çizilemez</div>';
  const bs=u.bantlar;
  const lo=Math.max(1,Math.floor((bs[3].ust||100)*0.90));
  const hi=Math.ceil(Math.max(bs[0].alt||0,u.buybox||0,u.buybox2||0,u.canli||0,u.oneri||0)*1.10);
  const W=1000,H=210,L=52,R=18,T=18,B=34;
  const X=f=>L+(Math.min(Math.max(f,lo),hi)-lo)/(hi-lo)*(W-L-R);
  const komAt=f=>kbKom(u,f), pctAt=f=>kbPct(u,f);
  // dikey ölçek
  let mn=1,mx=-1;const N=260;
  for(let i=0;i<=N;i++){const p=pctAt(lo+(hi-lo)*i/N);if(p==null)continue;mn=Math.min(mn,p);mx=Math.max(mx,p);}
  mn=Math.min(mn,0)-0.01; mx=Math.max(mx,u.minp)+0.015;
  const Y=p=>T+(mx-p)/(mx-mn)*(H-T-B);
  // bant bant kırık çizgi
  let sv=`<svg viewBox="0 0 ${W} ${H}" style="width:100%">`;
  sv+=`<line x1="${L}" y1="${Y(0)}" x2="${W-R}" y2="${Y(0)}" stroke="#cbd5e1" stroke-width="1"/>`;
  sv+=`<line x1="${L}" y1="${Y(u.minp)}" x2="${W-R}" y2="${Y(u.minp)}" stroke="#059669" stroke-width="1" stroke-dasharray="5 4"/>`;
  sv+=`<text x="${L-6}" y="${Y(u.minp)+4}" text-anchor="end" font-size="10" fill="#059669" font-weight="700">min %${(u.minp*100).toFixed(0)}</text>`;
  sv+=`<text x="${L-6}" y="${Y(0)+4}" text-anchor="end" font-size="10" fill="#94a3b8">0</text>`;
  const seg=[[lo,bs[3].ust],[bs[2].alt,bs[2].ust],[bs[1].alt,bs[1].ust],[bs[0].alt,hi]];
  for(const [a0,b0] of seg){
    if(a0==null||b0==null)continue;
    const a=Math.max(a0,lo), b=Math.min(b0,hi);
    if(b<=a)continue;
    let d='',ok=false;
    for(let i=0;i<=60;i++){const f=a+(b-a)*i/60;const p=pctAt(f);if(p==null)continue;
      d+=(ok?' L ':'M ')+X(f).toFixed(1)+' '+Y(p).toFixed(1);ok=true;}
    if(ok)sv+=`<path d="${d}" fill="none" stroke="#f27a1a" stroke-width="2.5" stroke-linecap="round"/>`;
    const km=komAt((a+b)/2);
    if(X(b)-X(a)>46)sv+=`<text x="${(X(a)+X(b))/2}" y="${H-16}" text-anchor="middle" font-size="10" fill="#697386">%${km}</text>`;}
  // buybox çizgisi
  if(u.buybox)sv+=`<line x1="${X(u.buybox)}" y1="${T}" x2="${X(u.buybox)}" y2="${H-B}" stroke="#dc2626" stroke-width="1.5" stroke-dasharray="4 3"/>
   <text x="${X(u.buybox)}" y="${T-4}" text-anchor="middle" font-size="10" font-weight="750" fill="#dc2626">buybox ${tk(u.buybox)}</text>`;
  if(u.buybox2)sv+=`<line x1="${X(u.buybox2)}" y1="${T}" x2="${X(u.buybox2)}" y2="${H-B}" stroke="#f472b6" stroke-width="1.2" stroke-dasharray="2 3"/>
   <text x="${X(u.buybox2)}" y="${T-4}" text-anchor="middle" font-size="10" font-weight="700" fill="#db2777">2. ${tk(u.buybox2)}</text>`;
  // en iyi satılabilir nokta
  const e=u.en_iyi;
  if(e){sv+=`<circle cx="${X(e.fiyat)}" cy="${Y(e.pct)}" r="6" fill="#059669" stroke="#fff" stroke-width="2"/>
   <text x="${X(e.fiyat)}" y="${Y(e.pct)-12}" text-anchor="middle" font-size="11" font-weight="800" fill="#059669">★ ${tk(e.fiyat)} · ${pc(e.pct)}</text>`;}
  if(u.canli){const p=pctAt(u.canli);
    if(p!=null)sv+=`<circle cx="${X(u.canli)}" cy="${Y(p)}" r="4.5" fill="#2563eb" stroke="#fff" stroke-width="2"/>
     <text x="${X(u.canli)}" y="${Y(p)+18}" text-anchor="middle" font-size="10" font-weight="700" fill="#2563eb">canlı ${tk(u.canli)}</text>`;}
  sv+=`<text x="${L}" y="${H-2}" font-size="10" fill="#697386">${tk(lo)} ₺</text>
   <text x="${W-R}" y="${H-2}" text-anchor="end" font-size="10" fill="#697386">${tk(hi)} ₺</text>`;
  // imleç takibi: şeffaf katman + nişangâh
  sv+=`<g id="kbcross" style="display:none;pointer-events:none">
   <line id="kbline" y1="${T}" y2="${H-B}" stroke="#101828" stroke-width="1" stroke-dasharray="3 3" opacity=".45"/>
   <circle id="kbdot" r="5" fill="#101828" stroke="#fff" stroke-width="2"/></g>
   <rect x="${L}" y="${T}" width="${W-L-R}" height="${H-T-B}" fill="transparent" style="cursor:crosshair"
    onmousemove="karHover(event)" onmouseleave="karHoverOff()"/>`;
  window._kb={lo,hi,L,R,T,B,W,H,mn,mx};
  return `<div id="kbwrap" style="position:relative">${sv}</svg>
   <div id="kbtip" style="position:absolute;display:none;pointer-events:none;background:#101828;color:#fff;
    padding:7px 11px;border-radius:9px;font-size:11px;line-height:1.7;white-space:nowrap;
    box-shadow:0 8px 24px rgba(16,24,40,.3);z-index:5"></div></div>`;
}

// ═══════════════ FIRSAT PROGRAMLARI (flaş + yıldızlı ürün) ═══════════════
let fAra='', fSuz='hepsi', fAcik=null;
async function firsatYukle(yenile){
  const c=document.getElementById('firsat');
  if(!window._fData||yenile) c.innerHTML='<div class="bos"><div><div class="big">⏳</div>Fırsat tarifeleri okunuyor...</div></div>';
  window._fData=JSON.parse(await window.pywebview.api.firsat_listesi(yenile?1:0));
  firsatCiz();
}
function fAraGir(el){const p=el.selectionStart;fAra=el.value.toLocaleLowerCase('tr');firsatCiz();
  const y=document.getElementById('fara');if(y){y.focus();y.setSelectionRange(p,p);}}
function fSuzGec(k){fSuz=k;firsatCiz();}
function fAc(bk){fAcik=(fAcik===bk?null:bk);firsatCiz();}
function firsatCiz(){
  const r=window._fData; if(!r) return;
  const o=r.ozet;
  const say=(f)=>r.urunler.filter(f).length;
  const iyi=x=>x.eniyi&&x.eniyi.pct>=0.10;
  let h=`<div class="kpis" style="margin-bottom:12px">
   <div class="kpi"><div class="t">Fırsat Teklifi</div><div class="v">${o.toplam}</div>
    <div class="d">${o.yildiz} yıldız · ${o.flas} flaş</div></div>
   <div class="kpi k-g"><div class="t">%10+ Kârla Girilebilir</div><div class="v">${say(iyi)}</div>
    <div class="d">en iyi seçenek üzerinden</div></div>
   <div class="kpi k-r"><div class="t">Her Seçenek Zararlı</div><div class="v">${say(x=>x.eniyi&&x.eniyi.pct<0)}</div>
    <div class="d">fırsata girme</div></div>
   <div class="kpi"><div class="t">Kaynak Dosyalar</div><div class="v" style="font-size:12px;line-height:1.5">
    ${r.kaynak.yildiz||'—'}<br>${r.kaynak.flas||'—'}</div><div class="d">Tarifeler/ arşivi</div></div>
  </div>
  <div style="display:flex;gap:8px;align-items:center;margin-bottom:12px;flex-wrap:wrap">
   <input id="fara" placeholder="🔍 ürün, barkod veya kategori ara" value="${fAra}" oninput="fAraGir(this)"
    style="padding:8px 12px;border:1px solid var(--line);border-radius:10px;width:260px;outline:none">
   <div class="seg">${[['hepsi','Hepsi'],['flas','⚡ Flaş'],['yildiz','★ Yıldız'],['iyi','%10+ kârlı'],['kotu','Zararlı']]
     .map(k=>`<div class="${fSuz===k[0]?'on':''}" onclick="fSuzGec('${k[0]}')">${k[1]}</div>`).join('')}</div>
   <span style="flex:1"></span>
   <button class="btn ghost" onclick="firsatYukle(1)">⟳ Tarifeleri yeniden oku</button></div>`;
  let L=r.urunler.filter(x=>{
    if(fSuz==='flas'&&!x.flas_var) return false;
    if(fSuz==='yildiz'&&!x.yildiz_var) return false;
    if(fSuz==='iyi'&&!iyi(x)) return false;
    if(fSuz==='kotu'&&!(x.eniyi&&x.eniyi.pct<0)) return false;
    if(fAra&&!((x.ad||'')+x.barkod+(x.kategori||'')).toLocaleLowerCase('tr').includes(fAra)) return false;
    return true;});
  h+=`<div class="card" style="padding:8px 16px"><table>
   <tr><th>Ürün</th><th style="text-align:right">Stok</th><th style="text-align:right">Mevcut</th>
   <th style="text-align:right">Kâr%</th><th style="text-align:right">En iyi seçenek</th>
   <th style="text-align:right">Fiyat</th><th style="text-align:right">Kâr%</th><th></th></tr>`;
  for(const x of L){
    const e=x.eniyi;
    const rz=[x.yildiz_var?'<span class="badge bg-w">★</span>':'', x.flas_var?`<span class="badge bg-r">⚡${x.slot}</span>`:''].join(' ');
    h+=`<tr style="cursor:pointer" onclick="fAc('${x.barkod}')">
     <td><b>${(x.ad||x.barkod).slice(0,46)}</b> ${rz}
      <div class="mut" style="font-size:10px">${x.barkod} · ${x.kategori||''}${x.maliyet!=null?' · maliyet '+tl(x.maliyet):' · <b style="color:#b45309">maliyet yok</b>'}</div></td>
     <td style="text-align:right">${x.stok??'—'}</td>
     <td style="text-align:right">${tl(x.fiyat)}</td>
     <td style="text-align:right" class="${x.simdi_pct==null?'mut':(x.simdi_pct>=0?'g':'r')}">${x.simdi_pct==null?'—':pc(x.simdi_pct)}</td>
     <td style="text-align:right"><b>${e?e.etiket:'—'}</b></td>
     <td style="text-align:right">${e?tl(e.fiyat):'—'}</td>
     <td style="text-align:right;font-weight:800" class="${!e?'mut':(e.pct>=0.10?'g':(e.pct>=0?'w':'r'))}">${e?pc(e.pct):'—'}</td>
     <td style="text-align:right" class="mut">${fAcik===x.barkod?'▾':'▸'}</td></tr>`;
    if(fAcik===x.barkod) h+=`<tr><td colspan="8" style="background:#fafbfd;padding:0 0 14px">${firsatDetay(x)}</td></tr>`;}
  h+='</table></div>';
  document.getElementById('firsat').innerHTML=h||'';
}
function firsatDetay(x){
  if(x.maliyet==null) return '<div style="padding:14px;color:#b45309"><b>Bu ürünün maliyeti girilmemiş</b> — Ürünler ekranından gir, kâr hesabı buraya akar.</div>';
  let h=`<div style="padding:12px 6px 0"><table style="width:100%">
   <tr><th>Seçenek</th><th style="text-align:right">Fiyat</th><th style="text-align:right">Komisyon</th>
   <th style="text-align:right">Stopaj %1</th><th style="text-align:right">Kargo (desi ${x.desi??'—'})</th>
   <th style="text-align:right">Hizmet</th><th style="text-align:right">Ürün maliyeti</th>
   <th style="text-align:right">NET KÂR</th><th style="text-align:right">Kâr%</th></tr>`;
  const en=x.eniyi?x.eniyi.etiket:null;
  for(const s of x.secenekler){
    if(s.eksik){h+=`<tr><td>${s.etiket}</td><td colspan="8" class="mut">hesaplanamadı</td></tr>`;continue;}
    const vur=s.etiket===en?'background:#ecfdf5;font-weight:700':'';
    const ind=(x.fiyat&&s.fiyat<x.fiyat)?`<span class="mut" style="font-size:10px"> −%${(100*(1-s.fiyat/x.fiyat)).toFixed(0)}</span>`:'';
    h+=`<tr style="${vur}">
     <td>${s.etiket===en?'★ ':''}${s.etiket}</td>
     <td style="text-align:right"><b>${tl(s.fiyat)}</b>${ind}</td>
     <td style="text-align:right" class="r">−${tl(s.komtl)}<span class="mut" style="font-size:10px"> %${s.kom}</span></td>
     <td style="text-align:right" class="r">−${tl(s.stopaj)}</td>
     <td style="text-align:right" class="r">−${tl(s.kargo)}</td>
     <td style="text-align:right" class="r">−${tl(s.platform)}</td>
     <td style="text-align:right" class="r">−${tl(s.maliyet)}</td>
     <td style="text-align:right;font-weight:850" class="${s.net>=0?'g':'r'}">${tl(s.net)}</td>
     <td style="text-align:right;font-weight:800" class="${s.pct>=0.10?'g':(s.pct>=0?'w':'r')}">${pc(s.pct)}</td></tr>`;}
  h+='</table>';
  if(x.tarihler&&x.tarihler.length) h+=`<div class="mut" style="font-size:11px;padding:8px 4px 0">⚡ flaş slotları: ${x.tarihler.join(' · ')}</div>`;
  h+=`<div class="mut" style="font-size:11px;padding:6px 4px 0">Komisyon oranı fiyatın düştüğü BAREME göre alınır — fiyat inince komisyon da düşer, o yüzden indirimin kâra darbesi göründüğünden hafiftir.</div></div>`;
  return h;
}

function karHover(e){
  const u=window._u, g=window._kb; if(!u||!g) return;
  const svg=document.querySelector('#kbwrap svg'); if(!svg) return;
  const r=svg.getBoundingClientRect(), sc=r.width/g.W;
  let xv=Math.min(Math.max((e.clientX-r.left)/sc, g.L), g.W-g.R);
  const f=Math.round(g.lo+(xv-g.L)/(g.W-g.L-g.R)*(g.hi-g.lo));
  const k=kbKom(u,f), kar=kbKar(u,f), p=kbPct(u,f);
  if(p==null){karHoverOff();return;}
  const Y=g.T+(g.mx-p)/(g.mx-g.mn)*(g.H-g.T-g.B);
  const c=document.getElementById('kbcross');
  c.style.display=''; 
  const ln=document.getElementById('kbline'); ln.setAttribute('x1',xv); ln.setAttribute('x2',xv);
  const dt=document.getElementById('kbdot'); dt.setAttribute('cx',xv); dt.setAttribute('cy',Y);
  const renk=p>=u.minp?'#34d399':(p>=0?'#fbbf24':'#f87171');
  const tip=document.getElementById('kbtip');
  tip.style.display='block';
  tip.innerHTML=`<b style="font-size:12px">${tl(f)}</b> &nbsp;<span style="opacity:.7">komisyon %${k}</span>
   <br>kâr <b style="color:${renk}">${tl(kar)}</b> · <b style="color:${renk}">${pc(p)}</b>
   ${p<u.minp?`<br><span style="opacity:.65">min %${(u.minp*100).toFixed(0)} altı</span>`:''}
   ${u.buybox?`<br><span style="opacity:.65">buybox ${f<=u.buybox-1?'ALINIR':'alınmaz'}</span>`:''}`;
  const tw=tip.offsetWidth||150;
  tip.style.left=Math.max(0,Math.min(r.width-tw, xv*sc+14))+'px';
  tip.style.top=Math.max(0,Y*sc-14)+'px';
}
function karHoverOff(){
  const c=document.getElementById('kbcross'), t=document.getElementById('kbtip');
  if(c)c.style.display='none'; if(t)t.style.display='none';
}
function bbKart(u){
  const bant=`<b>%${(u.minp*100).toFixed(0)} – %${(u.tavan*100).toFixed(0)}</b>`;
  const kut=(renk,cerceve,baslik,govde)=>`<div class="card" style="margin-bottom:14px;background:${renk};border-color:${cerceve}">
    <div style="font-size:13px;font-weight:800;color:${cerceve==='#a7f3d0'?'#047857':(cerceve==='#fde68a'?'#b45309':'#475569')}">${baslik}</div>
    <div style="font-size:12px;margin-top:6px;line-height:1.8">${govde}</div></div>`;
  if(u.bb_durum==='tek') return kut('#f8fafc','#e2e8f0','👤 TEK SATICI',
    `Bu üründe rakip yok — buybox yarışı da yok. Fiyat tamamen bizim elimizde,
     ${u.canli_pct!=null?`şu anki kâr <b>${pc(u.canli_pct)}</b> (kabul aralığı ${bant}).`:''}
     Aksiyon önerilmiyor.`);
  if(u.bb_durum==='veriyok') return kut('#f8fafc','#e2e8f0','BUYBOX VERİSİ YOK',
    'Bu ürün için buybox servisinden kayıt gelmedi. Ekranı yeniden açarak ya da Ürünler ekranındaki 🎯 Buybox Çek ile tazeleyin.');
  if(u.bb_durum==='maliyetsiz') return kut('#fffaeb','#fde68a','MALİYET GİRİLMEMİŞ',
    'Buybox bizde ama alış maliyeti girilmediği için kâr marjı ölçülemiyor — aşağıdaki girdi formundan girin.');
  if(u.bb_durum==='dusuk') return kut('#fffbeb','#fde68a','💸 KÂR DÜŞÜK — FİYAT YÜKSELT',
    `Buybox bizde (<b>${tl(u.canli)}</b>) ama kâr <b class="r">${pc(u.canli_pct)}</b> —
     ${u.kategori||'bu kategori'} kabul aralığı ${bant}, aralığın altındayız.
     <br>Hedef fiyat <b>${tl(u.hedef)}</b> ${u.kisit==='rakip'
       ?`— hedef kâra kadar çıkamıyoruz: üstteki satıcı <b>${tl(u.buybox2)}</b>, buybox'ı korumak için
         %1 altında duruyoruz.`
       :`— hedef kâr <b>%${(u.hedp*100).toFixed(0)}</b>'ın fiyatı${u.buybox2?` (üst fiyat ${tl(u.buybox2)}, yer var)`:''}.`}
     <br>Adet başına <b class="g">+${tl(u.kacan)}</b> kâr → yeni marj <b>${pc(u.oneri_pct)}</b>.`);
  if(u.bb_durum==='sikisik') return kut('#fffbeb','#fde68a','⚠️ KÂR DÜŞÜK · RAKİP YER BIRAKMIYOR',
    `Kâr <b class="r">${pc(u.canli_pct)}</b>, kabul aralığı ${bant} — aralığın altındayız,
     ama üstteki satıcı <b>${tl(u.buybox2)}</b>: fiyatı yükseltirsek buybox'ı kaybederiz.
     <br><span class="mut">Seçenekler: maliyeti düşürmek, ürünü bırakmak, ya da buybox'ı gözden çıkarıp
     ${tl(u.tavan_fiyat?Math.floor(u.tavan_fiyat):null)} civarına çıkmak.</span>`);
  if(u.bb_durum==='ustunde') return kut('#ecfdf5','#a7f3d0','🏆 BUYBOX BİZDE · KÂR TAVANIN ÜSTÜNDE',
    `Kâr <b>${pc(u.canli_pct)}</b>, ${u.kategori||'bu kategori'} tavanı <b>%${(u.tavan*100).toFixed(0)}</b>.
     Buybox yine bizde, aksiyon gerekmiyor — bilgi amaçlı.
     ${u.buybox2?`Üstteki satıcı ${tl(u.buybox2)}.`:''}`);
  if(u.bb_durum==='rakipte_ucuz') return kut('#fffbeb','#fde68a','⚔️ RAKİPTE · FİYATIMIZ ZATEN DAHA DÜŞÜK',
    `Fiyatımız <b>${tl(u.canli)}</b>, buybox sahibinin fiyatı <b>${tl(u.buybox)}</b> —
     yani <b>${tl(u.buybox-u.canli)}</b> daha ucuzuz ama buybox yine ${u.bb_sira}. sırada bizde değil.
     <br><b>Buybox sadece fiyatla belirlenmiyor</b>: satıcı puanı, teslimat hızı, iptal/iade oranı,
     kargo süresi de giriyor. Daha fazla indirmek buybox'ı getirmez, sadece kârı yer —
     bu yüzden fiyat önerisi verilmiyor.
     ${u.canli_pct!=null?`Şu anki kâr <b>${pc(u.canli_pct)}</b> (kabul %${(u.minp*100).toFixed(0)}–%${(u.tavan*100).toFixed(0)}).`:''}`);
  if(u.bb_durum==='bizde') return kut('#ecfdf5','#a7f3d0','🏆 BUYBOX BİZDE — FİYAT DOĞRU',
    `Fiyatımız <b>${tl(u.canli)}</b>, kâr <b>${pc(u.canli_pct)}</b> — ${u.kategori||'bu kategori'}
     kabul aralığı ${bant} içinde. ${u.buybox2?`Üstteki satıcı ${tl(u.buybox2)} (fark ${pc(u.fark)}),`:''}
     aksiyon gerekmiyor.
     <span class="mut">Aralığı değiştirmek için Excel → Hedefler → Min / Tavan Kâr %.</span>`);
  return '';
}
function satilirUyari(u){
  if(u.maliyet==null||!u.buybox||!u.oneri) return '';
  const e=u.en_iyi;
  if(u.satilabilir){
    if(!u.durum.includes('BAREM')) return '';
    return `<div class="card" style="margin-bottom:14px;background:#ecfdf5;border-color:#a7f3d0">
      <div style="font-size:13px;font-weight:800;color:#047857">✓ BAREM DÜŞÜRÜLDÜ — buybox alınıyor</div>
      <div style="font-size:12px;margin-top:6px;line-height:1.8">
      ${tl(u.oneri)} fiyatı bir alt bareme düşüyor → komisyon <b>%${e.kom}</b>,
      kâr <b>${tl(e.kar)}</b> (<b>${pc(e.pct)}</b>).
      ${e.pct<u.minp?`<br><span style="color:#b45309">Min hedef %${(u.minp*100).toFixed(0)} — ${((u.minp-e.pct)*100).toFixed(1)} puan altında,
      tolerans payı (${(TOL*100).toFixed(0)} puan) içinde kabul edildi.</span>`:''}</div></div>`;}
  return `<div class="card" style="margin-bottom:14px;background:#fef2f2;border-color:#fecaca">
    <div style="font-size:13px;font-weight:800;color:#b91c1c">⚠ ÖNERİ SATILABİLİR DEĞİL</div>
    <div style="font-size:12px;margin-top:6px;line-height:1.8">
    ${tl(u.oneri)} buybox'ın (${tl(u.buybox)}) üstünde — bu fiyatta satış beklenmez.
    ${e?`<br>Buybox altında ulaşılabilen en iyi: <b>${tl(e.fiyat)}</b> → komisyon %${e.kom},
    kâr <b>${tl(e.kar)}</b> (<b>${pc(e.pct)}</b>) — min hedefin ${((u.minp-e.pct)*100).toFixed(1)} puan altında,
    tolerans dışı.`:''}</div></div>`;
}
function simHesap(u,f){
  const kom=kbKom(u,f);          // tarifeli/tarifesiz fark etmez, bant listesini gezer
  if(u.maliyet==null||kom==null)return null;
  const kargo=f<=199.99?47.4:(f<=399.99?85.4:u.kargo_desi_tl);
  const komTL=f*kom/100, stopaj=f*STOPAJ_ORAN;
  const kar=f-komTL-stopaj-kargo-PLATFORM-u.maliyet;
  return{kom,kar,pct:kar/f,komTL,stopaj,kargo,platform:PLATFORM,maliyet:u.maliyet};
}
function detay(u){
  if(!u)return;
  const p=u.plus;
  const dk={dusuk:'k-w',sikisik:'k-w',bizde:'k-g',ustunde:'k-g',maliyetsiz:'k-w',tek:'k-p',veriyok:'k-w'}[u.bb_durum]
    ||(u.durum.startsWith('✓')?'k-g':(u.durum.startsWith('BUYBOX ALTI')?'k-r':'k-w'));
  const bbBaslik={tek:'Tek satıcı',bizde:'Buybox BİZDE',dusuk:'Buybox BİZDE',ustunde:'Buybox BİZDE',
                  sikisik:'Buybox BİZDE',maliyetsiz:'Buybox BİZDE',veriyok:'Buybox verisi yok'}[u.bb_durum]
    ||(u.bb_sira?u.bb_sira+'. sıradayız':'Buybox rakipte');
  let h=`<div class="card"><h2>${u.ad||u.barkod} · ${u.barkod}${u.tarifesiz?' · <span class="tz">tarife yok</span>':''}</h2>
  <div class="kpis">
  <div class="kpi ${dk}"><div class="t">${u.aksiyon?'Önerilen Fiyat':'Canlı Fiyat'}</div>
  <div class="v">${tl(u.aksiyon?u.oneri:u.canli)}</div><div class="d">${u.durum}</div></div>
  <div class="kpi k-g"><div class="t">${u.grup===1?'Kaçan Kâr / adet':'Şu anki Kâr'}</div>
  <div class="v">${u.grup===1?tl(u.kacan):(u.canli_pct!=null?pc(u.canli_pct):tl(u.oneri_kar))}</div>
  <div class="d">${u.grup===1?('kâr '+pc(u.canli_pct)+' → '+pc(u.oneri_pct))
    :('kabul %'+(u.minp*100).toFixed(0)+'–%'+(u.tavan*100).toFixed(0))}</div></div>
  <div class="kpi k-r"><div class="t">${bbBaslik}</div><div class="v">${tl(u.buybox)}</div>
  <div class="d">${(u.buybox2&&u.bb_sira===1)?('üst fiyat '+tk(u.buybox2)+' ₺'+(u.fark!=null?' · fark '+pc(u.fark):'')):
     (u.bb_durum==='tek'?'rakip yok':(u.bb_sira>1?u.bb_sira+'. sıradayız · biz '+tk(u.canli)+' ₺':(u.bb_f?'altı → '+pc(u.bb_pct)+' kâr':'—')))}</div></div>
  <div class="kpi k-p"><div class="t">Plus Teklifi</div><div class="v">${p?tl(p.limit):'—'}</div>
  <div class="d">${p?('%'+p.kom+' → '+tl(p.kar)+(p.fark==null?'':' · fark '+tl(p.fark))):'yok'}</div></div>
  </div></div>`;
  h+=bbKart(u);
  if(u.bb_durum==='rakipte') h+=satilirUyari(u);
  if(!u.tarifesiz){
    h+=`<div class="card"><h2>Fiyat Haritası — baremler & işaretçiler (Dönem ${S.donem})</h2>${fiyatEkseni(u)}</div>`;
    h+=`<div class="card"><h2>Kâr Basamakları <span style="text-transform:none;letter-spacing:0" class="mut">· bant sınırında komisyon zıplar, kâr düşer</span></h2>${karBasamak(u)}</div>`;
  }else{
    h+=`<div class="card"><h2>Komisyon</h2><div style="font-size:12px;line-height:1.8">
    Bu ürün haftalık tarife dosyasında YOK → komisyon tek oran: <b>${u.kom_tek!=null?'%'+u.kom_tek:'bilinmiyor'}</b>
    ${u.kom_tek!=null?'<span class="mut">(manuel giriş ya da Trendyol API)</span>':''}.
    <br><span class="mut">Barem merdiveni bilinmediği için "barem düşür" önerisi yapılamaz;
    buybox ve %1 fiyat kontrolü normal çalışır.</span></div></div>`;
  }
  const lo=u.sim_lo, hi=u.sim_hi;
  h+=`<div class="card"><h2>Fiyat Simülatörü</h2>
  <input type="range" id="sim" min="${lo}" max="${hi}" value="${u.oneri||u.canli||lo}" oninput="simGoster()">
  <div style="display:flex;justify-content:space-between;margin-top:10px;align-items:center">
  <span class="mut">${tl(lo)}</span><span id="simval"></span><span class="mut">${tl(hi)}</span></div>
  <div id="simdok" style="margin-top:14px;font-size:12px;border-top:1px solid var(--line);padding-top:10px"></div></div>`;
  h+=`<div class="card"><h2>Girdiler · kaydedince Excel'e işlenir</h2><div class="row2">
  <div><label>Maliyet</label><input type="number" id="i-m" value="${u.maliyet??''}"></div>
  <div><label>Desi</label><input type="number" id="i-d" value="${u.desi??''}"></div>
  <div><label>Buybox</label><input type="number" id="i-b" value="${u.buybox??''}"></div>
  <div><button class="btn ghost" onclick="kaydet('${u.barkod}')">💾 Kaydet</button></div>
  <div style="border-left:1px solid var(--line);padding-left:16px"><label>Gönderilecek Fiyat</label><input type="number" id="i-f" value="${u.oneri??u.canli??''}"></div>
  <div><button class="btn" onclick="gonder('${u.barkod}')">📤 Trendyol'a Gönder</button></div>
  </div><div class="mut" style="margin-top:10px;font-size:11px">Senkron: ${u.sync||'—'} · desi kargo ${tl(u.kargo_desi_tl)}</div></div>`;
  document.getElementById('sag').innerHTML=h;window._u=u;simGoster();
}
function simGoster(){const u=window._u;const f=+document.getElementById('sim').value;
  const r=simHesap(u,f);
  document.getElementById('simval').innerHTML=r?
    `${tl(f)} → <span class="mut">%${r.kom} kom</span> → <span class="${cls(r.pct,u.minp)}"><b>${tl(r.kar)}</b> (${pc(r.pct)})</span>`
    :tl(f)+' → maliyet gir';
  const d=document.getElementById('simdok');
  if(d) d.innerHTML=r?[['Satış fiyatı',f],['Komisyon (%'+r.kom+')',-r.komTL],['Stopaj (%1, KDV hariç matrah)',-r.stopaj],
      ['Kargo (desi '+(u.desi??'—')+')',-r.kargo],['Platform hizmet bedeli',-r.platform],['Ürün maliyeti',-r.maliyet]]
      .map(x=>`<div style="display:flex;justify-content:space-between;padding:3px 0">
        <span class="mut">${x[0]}</span><span style="font-variant-numeric:tabular-nums" class="${x[1]<0?'r':''}">${tl(x[1])}</span></div>`).join('')
      +`<div style="display:flex;justify-content:space-between;border-top:1px solid var(--line);margin-top:6px;padding-top:6px;font-weight:800">
        <span>Net kâr</span><span class="${r.kar>=0?'g':'r'}">${tl(r.kar)} · ${pc(r.pct)}</span></div>`:'';}
function modalGoster(html,onay){
  const m=document.createElement('div');
  m.id='modal';m.style.cssText='position:fixed;inset:0;background:#10182866;display:grid;place-items:center;z-index:99';
  m.innerHTML=`<div style="background:#fff;border-radius:18px;padding:24px;max-width:440px;box-shadow:0 24px 70px rgba(16,24,40,.35)">
  ${html}<div style="display:flex;gap:10px;justify-content:flex-end;margin-top:18px">
  <button class="btn ghost" onclick="document.getElementById('modal').remove()">Vazgeç</button>
  ${onay?'<button class="btn" id="m-onay">Onayla ve Gönder</button>':''}</div></div>`;
  document.body.appendChild(m);
  if(onay)document.getElementById('m-onay').onclick=onay;
}
async function gonder(bk){
  const u=S.urunler.find(x=>x.barkod===bk);
  const f=parseFloat(document.getElementById('i-f').value);
  if(!f||f<=0){modalGoster('<b>Geçerli bir fiyat gir.</b>');return;}
  const sim=simHesap(u,f);
  modalGoster(`<h2 style="margin-bottom:12px;font-size:14px">${u.ad.slice(0,50)}</h2>
  <div style="font-size:13px;line-height:2">
  Mevcut fiyat: <b>${tl(u.canli)}</b><br>
  YENİ FİYAT: <b style="font-size:17px;color:var(--or)">${tl(f)}</b><br>
  ${sim?('Komisyon: <b>%'+sim.kom+'</b> · Beklenen kâr: <b class="'+cls(sim.pct,u.minp)+'">'+tl(sim.kar)+' ('+pc(sim.pct)+')</b><br>'):''}
  <span class="mut">Bu fiyat Trendyol'da CANLI olarak güncellenecek.</span></div>`,
  async ()=>{
    document.getElementById('m-onay').textContent='⏳ Gönderiliyor...';
    const r=JSON.parse(await window.pywebview.api.fiyat_gonder(bk,f));
    document.getElementById('modal').remove();
    modalGoster(`<b style="color:${r.ok?'var(--yes)':'var(--no)'}">${r.ok?'✓ Gönderildi':'✗ Gönderilemedi'}</b><div class="mut" style="margin-top:8px;font-size:12px">${r.mesaj}</div>`);
  });
}
async function kaydet(bk){const g=id=>document.getElementById(id).value;
  S=JSON.parse(await window.pywebview.api.kaydet(bk,g('i-m'),g('i-d'),g('i-b')));ciz();}
document.getElementById('donem').onclick=async e=>{if(e.target.dataset.d){S=JSON.parse(await window.pywebview.api.donem_sec(e.target.dataset.d));ciz();}};
document.querySelector('.rail').onclick=e=>{const a=e.target.closest('a');if(a&&a.dataset.g)gorunumSec(a.dataset.g);};
const BASLIK={karlilik:'Trendyol · Buybox & Fiyat Kontrol',ciro:'Trendyol · Ciro Raporu',urunler:'Trendyol · Ürünler',
 firsat:'Trendyol · Fırsat Programları (Flaş & Yıldız)',
 hb_ciro:'Hepsiburada · Ciro Raporu',hb_urunler:'Hepsiburada · Ürünler'};
function gorunumSec(g){
  document.querySelectorAll('.rail a').forEach(x=>x.classList.toggle('on',x.dataset.g===g));
  document.getElementById('baslik').textContent=BASLIK[g]||'';
  const hb=g.startsWith('hb_');
  document.getElementById('sol').style.display=g==='karlilik'?'flex':'none';
  document.getElementById('sag').style.display=g==='karlilik'?'block':'none';
  document.getElementById('ciro').style.display=g==='ciro'?'block':'none';
  document.getElementById('urunler').style.display=g==='urunler'?'block':'none';
  document.getElementById('firsat').style.display=g==='firsat'?'block':'none';
  document.getElementById('hb').style.display=hb?'block':'none';
  document.getElementById('donem').style.display=g==='karlilik'?'flex':'none';
  document.getElementById('gbtn').style.display=hb?'none':'inline-block';
  if(g==='karlilik') buyboxOto();
  if(g==='ciro'&&!window._ciroData) ciroYukle(30);
  if(g==='urunler'&&!window._uData) urunYukle(0);
  if(g==='firsat'&&!window._fData) firsatYukle(0);
  if(g==='hb_ciro'&&!window._hbCiro) hbCiroYukle(30);
  if(g==='hb_urunler'&&!window._hbU) hbUrunYukle(0);
  if(hb) hbCiz(g);
}
async function hbCiroYukle(gun){
  const c=document.getElementById('hb');
  c.innerHTML='<div class="bos"><div><div class="big">⏳</div>Hepsiburada verisi çekiliyor...</div></div>';
  window._hbCiro=JSON.parse(await window.pywebview.api.hb_ciro(gun));
  hbCiz('hb_ciro');
}
async function hbUrunYukle(y){
  const c=document.getElementById('hb');
  c.innerHTML='<div class="bos"><div><div class="big">⏳</div>Hepsiburada ürünleri çekiliyor (~30 sn)...</div></div>';
  window._hbU=JSON.parse(await window.pywebview.api.hb_urun_listesi(y));
  hbCiz('hb_urunler');
}
let hbAra='', hbF='is';
function hbAraGir(el){const p=el.selectionStart;hbAra=el.value.toLocaleLowerCase('tr');hbCiz('hb_urunler');
  const y=document.getElementById('hbara');if(y){y.focus();y.setSelectionRange(p,p);}}
function hbCiz(g){
  if(g==='hb_ciro'&&window._hbCiro){hbCiroCiz();return;}
  if(g==='hb_urunler'&&window._hbU){hbUrunCiz();return;}
  document.getElementById('hb').innerHTML=`
  <div class="card" style="max-width:620px">
  <h2>Hepsiburada bağlantısı</h2>
  <div style="font-size:13px;line-height:1.9">
  Bu ekran <b>${g==='hb_ciro'?'ciro & net kâr raporunu':'ürün/maliyet listesini'}</b> Hepsiburada verisiyle
  Trendyol'daki yapının aynısı olarak gösterecek. Bağlamak için API bilgileri gerekiyor:
  <ol style="margin:10px 0 0 18px;line-height:2">
  <li>HB Satıcı Paneli → <b>Hesabım → Entegrasyon Bilgileri → API Anahtarı</b></li>
  <li>Oradaki <b>kullanıcı adı</b>, <b>şifre</b> ve <b>Merchant ID</b>'yi Claude'a ver</li>
  </ol>
  <div class="mut" style="margin-top:12px;font-size:12px">
  Not: HB'de komisyon ürün bazlı barem değil, <b>kategori bazlı</b>. Bu yüzden HB tarafında "Tarife Takibi"
  yerine kategori komisyonu + fiyat/kâr simülasyonu olacak. Maliyetler iki pazaryerinde <b>ortak</b>
  kullanılacak — aynı ürün için "Trendyol'da mı HB'de mi daha kârlı" karşılaştırması çıkacak.</div>
  </div></div>`;
}
let uFiltre='is', uAra='';
async function urunYukle(yenile){
  const c=document.getElementById('urunler');
  if(!window._uData||yenile) c.innerHTML='<div class="bos"><div><div class="big">⏳</div>Trendyol ürünleri çekiliyor (~20-40 sn)...</div></div>';
  window._uData=JSON.parse(await window.pywebview.api.urun_listesi(yenile));
  urunCiz();
}
function uAraGir(el){
  const p=el.selectionStart; uAra=el.value.toLocaleLowerCase('tr'); urunCiz();
  const y=document.getElementById('uara'); if(y){y.focus(); y.setSelectionRange(p,p);}
}
function urunCiz(){
  const D=window._uData, o=D.ozet;
  let liste=D.urunler.filter(x=>{
    if(uFiltre==='is'&&(x.maliyet!=null||!(x.stok>0))) return false;
    if(uFiltre==='eksik'&&x.maliyet!=null) return false;
    if(uFiltre==='stok'&&!(x.stok>0)) return false;
    if(uFiltre==='maliyetli'&&x.maliyet==null) return false;
    if(uAra&&!(x.ad+x.barkod+x.kategori+x.marka).toLocaleLowerCase('tr').includes(uAra)) return false;
    return true;});
  const f=(k,l)=>`<div class="${uFiltre===k?'on':''}" onclick="uFiltre='${k}';urunCiz()">${l}</div>`;
  let h=`<div style="display:flex;gap:10px;align-items:center;margin-bottom:14px">
  <h2 style="font-size:15px;font-weight:800">Ürünler <span class="mut" style="font-weight:500;font-size:12px">· ${o.toplam} satıştaki ürün · ${o.maliyetli} maliyetli</span></h2>
  <span style="flex:1"></span>
  <input id="uara" placeholder="🔍 isim, barkod, marka, kategori..." value="${uAra}" oninput="uAraGir(this)"
   style="padding:8px 12px;border:1px solid var(--line);border-radius:10px;width:260px;outline:none">
  <div class="seg">${f('is','⚡ Stokta & maliyetsiz')}${f('stok','Stokta')}${f('eksik','Maliyeti eksik')}${f('maliyetli','Maliyetli')}${f('hepsi','Hepsi')}</div>
  <button class="btn ghost" onclick="urunYukle(1)">⟳ Yenile</button>
  <button class="btn ghost" id="bbbtn" onclick="buyboxCek()">🎯 Buybox Çek</button>
  <button class="btn" id="tkbtn" onclick="topluKaydet()" disabled>💾 Kaydet</button></div>
  <div class="kpis" style="margin-bottom:14px">
  <div class="kpi"><div class="t">Satıştaki Ürün</div><div class="v">${o.toplam}</div><div class="d">${o.stokta} tanesi stokta</div></div>
  <div class="kpi k-g"><div class="t">Maliyeti Girilmiş</div><div class="v">${o.maliyetli}</div><div class="d">${pc(o.maliyetli/o.toplam)} tamamlandı</div></div>
  <div class="kpi k-w"><div class="t">Maliyet Bekleyen</div><div class="v">${o.toplam-o.maliyetli}</div><div class="d">net kâra katılamaz</div></div>
  <div class="kpi k-r"><div class="t">Komisyonu Bilinmeyen</div><div class="v">${o.komsuz}</div><div class="d">${o.kom_manuel} tanesi manuel girilmiş</div></div></div>
  <div class="card" style="padding:8px 16px"><table><tr>
  <th>Ürün</th><th style="text-align:right">Fiyat</th><th style="text-align:right">Stok</th>
  <th style="text-align:right">Komisyon</th><th style="text-align:right">Birim Maliyet</th><th style="text-align:right">Desi</th></tr>`;
  for(const x of liste.slice(0,400)){
    const eks=x.maliyet==null;
    h+=`<tr data-eks="${eks?1:0}" ${eks?'style="background:#fffbeb"':''}>
    <td><b>${(x.ad||x.barkod).slice(0,50)}</b><div class="mut" style="font-size:10px">${x.barkod}</div></td>
    <td style="text-align:right">${tl(x.fiyat)}</td>
    <td style="text-align:right" class="${x.stok>0?'':'mut'}">${x.stok??'—'}</td>
    <td style="text-align:right">${x.kom?('<b>%'+x.kom+'</b><div class="mut" style="font-size:9px">'+x.kom_kaynak+'</div>'):'<span class="mut">—</span>'}</td>
    <td style="text-align:right"><input type="number" id="um-${x.barkod}" value="${x.maliyet??''}" placeholder="gir" data-o="${x.maliyet??''}"
      style="width:100px;padding:5px 8px;text-align:right" oninput="degisti(this)" onkeydown="uEnter(event,this)"></td>
    <td style="text-align:right"><input type="number" id="ud-${x.barkod}" value="${x.desi??''}" placeholder="—" data-o="${x.desi??''}"
      style="width:64px;padding:5px 8px;text-align:right" oninput="degisti(this)" onkeydown="uEnter(event,this)"></td></tr>`;}
  h+='</table>'+(liste.length>400?'<div class="mut" style="padding:10px">İlk 400 satır gösteriliyor — aramayı daralt.</div>':'')+'</div>';
  document.getElementById('urunler').innerHTML=h;
}
function degisti(el){
  el.style.borderColor = (el.value!==el.dataset.o) ? 'var(--or)' : 'var(--line)';
  el.closest('tr').style.background = bekleyen(el.closest('tr')) ? '#fff7ed' : (el.closest('tr').dataset.eks==='1'?'#fffbeb':'');
  sayacGuncelle();
}
function bekleyen(tr){return [...tr.querySelectorAll('input')].some(i=>i.value!==i.dataset.o);}
function bekleyenler(){
  const out=[];
  document.querySelectorAll('#urunler tr, #hb tr').forEach(tr=>{
    const m=tr.querySelector('input[id^=um-]'), d=tr.querySelector('input[id^=ud-]'), k=tr.querySelector('input[id^=uk-]');
    if(!m) return;
    if(m.value!==m.dataset.o || d.value!==d.dataset.o || (k&&k.value!==k.dataset.o))
      out.push({barkod:m.id.slice(3), maliyet:m.value, desi:d.value, kom:k?k.value:''});
  });
  return out;
}
function sayacGuncelle(){
  const n=bekleyenler().length;
  for(const id of ['tkbtn','hbkbtn']){const b=document.getElementById(id);
    if(b){b.disabled=n===0; b.textContent=n?`💾 Kaydet (${n})`:'💾 Kaydet';}}
}
function uEnter(e,el){
  if(e.key!=='Enter')return;
  const hepsi=[...document.querySelectorAll('#urunler input[type=number]')];
  const i=hepsi.indexOf(el);
  if(e.metaKey||e.ctrlKey){topluKaydet();return;}
  if(hepsi[i+1]){hepsi[i+1].focus();hepsi[i+1].select();}
}
async function topluKaydet(){
  const kay=bekleyenler(); if(!kay.length)return;
  const hbEkran=document.getElementById('hb').style.display!=='none';
  const b=document.getElementById(hbEkran?'hbkbtn':'tkbtn');
  if(b){b.textContent='⏳ Kaydediliyor...'; b.disabled=true;}
  try{
    const res=JSON.parse(await window.pywebview.api.toplu_kaydet(JSON.stringify(kay)));
    if(res.urunler&&res.ozet) window._uData=res;
    if(res.hb) window._hbU=res.hb;
    window._ciroData=null; window._hbCiro=null;
    S=JSON.parse(await window.pywebview.api.durum());
    if(hbEkran) hbUrunCiz(); else urunCiz();
    const n=(res.kaydedilen!=null)?res.kaydedilen:kay.length;
    modalGoster(`<b style="color:var(--yes)">✓ ${n} ürün kaydedildi</b>
    <div class="mut" style="margin-top:8px;font-size:12px">Ciro raporları güncellendi.</div>`);
  }catch(e){
    if(b){b.textContent='💾 Kaydet'; b.disabled=false;}
    modalGoster(`<b style="color:var(--no)">✗ Kaydedilemedi</b><div class="mut" style="margin-top:8px;font-size:12px">${e}</div>`);
  }
}
async function buyboxOto(){
  // Ekran her açıldığında STOKTAKİ tüm ürünlerin buybox'ı tazelenir (tarifede olsun olmasın) —
  // hem sıramız hem 2. satıcı fiyatı güncel olsun diye. ~74 ürün = 8 istek.
  const d=document.getElementById('bbdurum'); if(!d) return;
  d.textContent='🎯 buybox kontrol ediliyor...';
  try{
    const r=JSON.parse(await window.pywebview.api.buybox_cek(1));
    S=r.durum; ciz();
    const st=new Date().toLocaleTimeString('tr-TR',{hour:'2-digit',minute:'2-digit'});
    const say=g=>S.urunler.filter(u=>u.grup===g).length;
    d.innerHTML=`🎯 ${st} · ${S.urunler.length} stoklu ürün · `
      +`<b class="r">${say(1)}</b> kâr düşük · <b>${say(2)+say(3)}</b> rakipte/eksik · `
      +`${say(4)} doğru/tavanda · ${say(5)} tek satıcı`;
  }catch(e){ d.textContent='🎯 buybox çekilemedi: '+e; }
}
async function buyboxCek(){
  const b=document.getElementById('bbbtn'); b.textContent='⏳ Buybox çekiliyor...'; b.disabled=true;
  try{
    const r=JSON.parse(await window.pywebview.api.buybox_cek());
    S=r.durum; window._ciroData=null;
    window._uData=JSON.parse(await window.pywebview.api.urun_listesi(0));
    urunCiz();
    modalGoster(`<b style="color:var(--yes)">✓ Buybox güncellendi</b>
    <div class="mut" style="margin-top:8px;font-size:12px">${r.cekilen} üründe buybox fiyatı bulundu, ${r.yazilan} kayıt Excel'e yazıldı.
    Tarife Takibi ekranındaki öneriler artık güncel buybox'a göre.</div>`);
  }catch(e){ modalGoster(`<b style="color:var(--no)">✗ Hata</b><div class="mut">${e}</div>`); }
  b.textContent='🎯 Buybox Çek'; b.disabled=false;
}
async function urunKaydet(bk){
  const m=document.getElementById('um-'+bk).value, d=document.getElementById('ud-'+bk).value;
  const btn=event&&event.target&&event.target.tagName==='BUTTON'?event.target:null;
  if(btn){btn.textContent='✓';btn.disabled=true;}
  window._uData=JSON.parse(await window.pywebview.api.urun_kaydet(bk,m,d));
  window._ciroData=null;                       // ciro ekranı tazelensin
  S=JSON.parse(await window.pywebview.api.durum());
  urunCiz();
}
async function malKaydet(bk){
  const el=document.getElementById('mi-'+bk); const v=parseFloat((el.value||'').replace(',','.'));
  if(!v||v<=0){el.style.borderColor='#dc2626';return;}
  el.disabled=true;
  window._ciroData=JSON.parse(await window.pywebview.api.maliyet_kaydet(bk,v));
  ciroCiz();
  S=JSON.parse(await window.pywebview.api.durum());
}
function hbCiroCiz(){
  const r=window._hbCiro, t=r.toplam;
  let h=`<div style="display:flex;gap:10px;align-items:center;margin-bottom:14px">
  <h2 style="font-size:15px;font-weight:800">Hepsiburada · son ${r.gun} gün
  <span class="mut" style="font-weight:500;font-size:12px">· yerel defter: ${r.defter} paket${r.yeni?' (+'+r.yeni+' yeni)':''}</span></h2>
  <span style="flex:1"></span>
  <div class="seg">${[7,14,30,60].map(g=>`<div class="${g===r.gun?'on':''}" onclick="hbCiroYukle(${g})">${g}g</div>`).join('')}</div>
  <button class="btn ghost" onclick="gizliDegis2()">👁 ${GIZLI?'Göster':'Gizle'}</button></div>
  <div class="kpis" style="margin-bottom:10px">
  <div class="kpi"><div class="t">Brüt Ciro</div><div class="v">${tlm(t.ciro)}</div><div class="d">${t.adet} adet</div></div>
  <div class="kpi k-r"><div class="t">Komisyon + Kesintiler</div><div class="v">${tlm(t.kom_all)}</div><div class="d">kom ${tlm(t.kom)} (KDV'li) + %1 stopaj ${tlm(t.stopaj)} + %0,8 tahsilat ${tlm(t.tahsilat)}${t.indirim>0.01?' − HB indirimi '+tlm(t.indirim):''}</div></div>
  <div class="kpi k-w"><div class="t">Kargo + hizmet bedeli</div><div class="v">${tlm(t.kargo)}</div><div class="d">${t.ciro?pc(t.kargo/t.ciro):'—'} / ciro</div></div>
  <div class="kpi"><div class="t">Ürün Maliyeti</div><div class="v">${tlm(t.maliyet)}</div><div class="d">${t.ciro?pc(t.maliyet/t.ciro):'—'} / ciro</div></div></div>
  <div class="card" style="margin-bottom:14px;background:linear-gradient(135deg,#c2410c,#ea580c);border:0">
  <div style="display:flex;align-items:center;gap:20px;color:#fff">
  <div><div style="font-size:11px;opacity:.85;text-transform:uppercase;letter-spacing:1px;font-weight:750">Net Kâr</div>
  <div style="font-size:30px;font-weight:900;margin-top:2px">${tlm(t.net)}</div></div>
  <div style="border-left:1px solid #ffffff40;padding-left:20px"><div style="font-size:11px;opacity:.85">Net marj</div>
  <div style="font-size:22px;font-weight:850">${pc(t.marj)}</div></div>
  <div style="flex:1;font-size:11px;opacity:.9;line-height:1.7">ciro − komisyon(KDV'li) − %1 stopaj − %0,8 tahsilat + HB indirimi − kargo − ürün maliyeti
  ${r.eksik.length?('<br><b>⚠ '+r.eksik.length+' üründe maliyet yok</b>'):''}</div></div></div>
  <div class="card" style="padding:8px 16px"><table>
  <tr><th>Ürün</th><th style="text-align:right">Adet</th><th style="text-align:right">Ciro</th>
  <th style="text-align:right">Kom+Kesinti</th><th style="text-align:right">Kargo+Hizmet</th>
  <th style="text-align:right">Maliyet</th><th style="text-align:right">Net Kâr</th><th style="text-align:right">Marj</th></tr>`;
  for(const x of r.satirlar){
    h+=`<tr ${x.maliyet==null?'style="background:#fffbeb"':''}><td><b>${(x.ad||x.barkod).slice(0,44)}</b>
    <div class="mut" style="font-size:10px">${x.barkod}</div></td>
    <td style="text-align:right">${x.adet}</td><td style="text-align:right;font-weight:750">${tlm(x.ciro)}</td>
    <td style="text-align:right" class="r">${tlm(x.kom_all)}</td><td style="text-align:right" class="w">${tlm(x.kargo)}</td>
    <td style="text-align:right" class="mut">${x.maliyet==null?'—':tlm(x.maliyet)}</td>
    <td style="text-align:right;font-weight:850" class="${x.net==null?'mut':(x.net>=0?'g':'r')}">${x.net==null?'maliyet gir':tlm(x.net)}</td>
    <td style="text-align:right;font-weight:750" class="${x.marj==null?'mut':(x.marj>=0?'g':'r')}">${pc(x.marj)}</td></tr>`;}
  document.getElementById('hb').innerHTML=h+'</table></div>';
}
function gizliDegis2(){GIZLI=!GIZLI;hbCiroCiz();}
function hbUrunCiz(){
  const D=window._hbU, o=D.ozet;
  const liste=D.urunler.filter(x=>{
    if(hbF==='is'&&(x.maliyet!=null||!(x.stok>0))) return false;
    if(hbF==='stok'&&!(x.stok>0)) return false;
    if(hbF==='maliyetli'&&x.maliyet==null) return false;
    if(hbAra&&!((x.ad||'')+x.barkod+x.sku+(x.msku||'')+x.marka).toLocaleLowerCase('tr').includes(hbAra)) return false;
    return true;});
  const f=(k,l)=>`<div class="${hbF===k?'on':''}" onclick="hbF='${k}';hbCiz('hb_urunler')">${l}</div>`;
  let h=`<div style="display:flex;gap:10px;align-items:center;margin-bottom:14px">
  <h2 style="font-size:15px;font-weight:800">Hepsiburada · Ürünler <span class="mut" style="font-weight:500;font-size:12px">· ${o.toplam} ürün · ${o.maliyetli} maliyetli</span></h2>
  <span style="flex:1"></span>
  <input id="hbara" placeholder="🔍 ara..." value="${hbAra}" oninput="hbAraGir(this)"
   style="padding:8px 12px;border:1px solid var(--line);border-radius:10px;width:240px;outline:none">
  <div class="seg">${f('is','⚡ Stokta & maliyetsiz')}${f('stok','Stokta')}${f('maliyetli','Maliyetli')}${f('hepsi','Hepsi')}</div>
  <button class="btn" id="hbkbtn" onclick="topluKaydet()" disabled>💾 Kaydet</button></div>
  <div class="kpis" style="margin-bottom:14px">
  <div class="kpi"><div class="t">HB Ürünü</div><div class="v">${o.toplam}</div><div class="d">${o.stokta} stokta</div></div>
  <div class="kpi k-g"><div class="t">Maliyetli</div><div class="v">${o.maliyetli}</div><div class="d">${pc(o.maliyetli/o.toplam)}</div></div>
  <div class="kpi k-w"><div class="t">Maliyet Bekleyen</div><div class="v">${o.toplam-o.maliyetli}</div><div class="d">ortak maliyet havuzu</div></div>
  <div class="kpi"><div class="t">Komisyon Oranı Bilinen</div><div class="v">${o.komlu}</div><div class="d">${o.kampanyali} üründe aktif kampanya</div></div></div>
  <div class="card" style="padding:8px 16px"><table><tr>
  <th>Ürün</th><th style="text-align:right">Fiyat</th><th style="text-align:right">Stok</th>
  <th style="text-align:right">Komisyon</th><th style="text-align:right">Kampanya</th>
  <th style="text-align:right">Birim Maliyet</th><th style="text-align:right">Desi</th></tr>`;
  for(const x of liste.slice(0,400)){
    const eks=x.maliyet==null;
    h+=`<tr data-eks="${eks?1:0}" ${eks?'style="background:#fffbeb"':''}>
    <td><b>${(x.ad||x.sku).slice(0,50)}</b><div class="mut" style="font-size:10px">${x.barkod||x.sku}${x.katalogda?'':' · listing'}</div></td>
    <td style="text-align:right">${tl(x.fiyat)}</td><td style="text-align:right" class="${x.stok>0?'':'mut'}">${x.stok??'—'}</td>
    <td style="text-align:right">${x.kom!=null?('<b>%'+x.kom+'</b>'+(x.vade?'<div class="mut" style="font-size:9px">'+x.vade+' gün vade</div>':'')):'<span class="mut">—</span>'}</td>
    <td style="text-align:right">${x.kamp_fiyat?('<b class="w">'+tl(x.kamp_fiyat)+'</b><div class="mut" style="font-size:9px">'+(x.kamp_hb?'HB '+tl(x.kamp_hb)+' · ':'')+(x.kamp_bitis||'')+'</div>'):'<span class="mut">—</span>'}</td>
    <td style="text-align:right"><input type="number" id="um-${x.anahtar}" value="${x.maliyet??''}" placeholder="gir" data-o="${x.maliyet??''}"
      style="width:100px;padding:5px 8px;text-align:right" oninput="degisti(this)" onkeydown="uEnter(event,this)"></td>
    <td style="text-align:right"><input type="number" id="ud-${x.anahtar}" value="${x.desi??''}" placeholder="—" data-o="${x.desi??''}"
      style="width:64px;padding:5px 8px;text-align:right" oninput="degisti(this)" onkeydown="uEnter(event,this)"></td></tr>`;}
  document.getElementById('hb').innerHTML=h+'</table></div>';
  sayacGuncelle();
}
async function ciroYukle(gun){
  const c=document.getElementById('ciro');
  c.innerHTML='<div class="bos"><div><div class="big">⏳</div>Rapor hazırlanıyor — API taranıyor (~20-50 sn)...</div></div>';
  try{window._ciroData=JSON.parse(await window.pywebview.api.ciro(gun));ciroCiz();}
  catch(e){c.innerHTML='<div class="bos"><div>Hata: '+e+'</div></div>';}
}
function ciroCiz(){
  const r=window._ciroData, t=r.toplam;
  const maxC=Math.max(...r.satirlar.map(x=>x.ciro),1);
  let bars='';
  for(const x of r.satirlar.filter(y=>!cAra||((y.ad||'')+y.barkod).toLocaleLowerCase('tr').includes(cAra)).slice(0,8)){
    const w=v=>Math.max(0,v/maxC*100);
    bars+=`<div style="margin-bottom:11px"><div style="display:flex;justify-content:space-between;font-size:11.5px;margin-bottom:4px">
    <b>${(x.ad||x.barkod).slice(0,44)}</b><span><b>${tlm(x.ciro)}</b> <span class="mut">· net <b class="${x.net==null?'mut':(x.net>=0?'g':'r')}">${x.net==null?'maliyet yok':tlm(x.net)}</b></span></span></div>
    <div style="display:flex;height:16px;border-radius:6px;overflow:hidden;background:#f0f2f7">
    <div style="width:${w(x.net||0)}%;background:linear-gradient(90deg,#10b981,#059669)"></div>
    <div style="width:${w(x.maliyet||0)}%;background:#94a3b8"></div>
    <div style="width:${w(x.kom_all)}%;background:#f87171"></div>
    <div style="width:${w(x.kargo+(x.platform||0))}%;background:#fbbf24"></div></div></div>`;}
  let h=`<div style="display:flex;gap:10px;align-items:center;margin-bottom:14px">
  <h2 style="font-size:15px;font-weight:800">Son ${r.gun} gün <span class="mut" style="font-weight:500;font-size:12px">· sipariş tarihi bazlı</span></h2>
  <span style="flex:1"></span>
  <input id="cara" placeholder="🔍 ürün ara..." value="${cAra}" oninput="cAraGir(this)"
   style="padding:8px 12px;border:1px solid var(--line);border-radius:10px;width:220px;outline:none;margin-right:8px">
  <div class="seg">${[7,14,30,60].map(g=>`<div class="${g===r.gun?'on':''}" onclick="ciroYukle(${g})">${g}g</div>`).join('')}</div>
  <button class="btn ghost" onclick="gizliDegis()" title="Tutarları gizle/göster" style="padding:8px 12px">${GIZLI?'👁 Tutarları göster':'👁 Tutarları gizle'}</button></div>
  <div class="kpis" style="margin-bottom:10px">
  <div class="kpi"><div class="t">Brüt Ciro</div><div class="v">${tlm(t.ciro)}</div><div class="d">${t.adet} adet (iade düşülmüş)</div></div>
  <div class="kpi k-r"><div class="t">Komisyon + %1 Stopaj</div><div class="v">${tlm(t.kom_all)}</div><div class="d">kom ${tlm(t.kom)} + stopaj ${tlm(t.stopaj)}
  <br><span style="font-size:10px">${tlm(t.kom-(t.kom_o||0)-(t.kom_t||0))} hakediş · ${tlm(t.kom_o||0)} sipariş oranı · ${tlm(t.kom_t||0)} tahmin</span></div></div>
  <div class="kpi k-w"><div class="t">Kargo + Hizmet Bedeli</div><div class="v">${tlm(t.kargo+(t.platform||0))}</div><div class="d">kargo ${tlm(t.kargo)}${t.kargo_t>0.01?' ≈':''} + platform ${tlm(t.platform||0)}</div></div>
  <div class="kpi"><div class="t">Ürün Maliyeti</div><div class="v">${tlm(t.maliyet)}</div><div class="d">${t.ciro?pc(t.maliyet/t.ciro):'—'} / ciro</div></div>
  </div>
  <div class="card" style="margin-bottom:14px;background:linear-gradient(135deg,#0f766e,#059669);border:0">
  <div style="display:flex;align-items:center;gap:20px;color:#fff">
  <div><div style="font-size:11px;opacity:.85;text-transform:uppercase;letter-spacing:1px;font-weight:750">Net Kâr</div>
  <div style="font-size:30px;font-weight:900;margin-top:2px">${tlm(t.net)}</div></div>
  <div style="border-left:1px solid #ffffff40;padding-left:20px"><div style="font-size:11px;opacity:.85">Net marj</div>
  <div style="font-size:22px;font-weight:850">${pc(t.marj)}</div></div>
  <div style="flex:1;font-size:11px;opacity:.9;line-height:1.7">ciro − komisyon − %1 stopaj − kargo − platform hizmet bedeli (13,19 ₺/gönderi) − ürün maliyeti
  ${r.eksik.length?('<br><b>⚠ '+r.eksik.length+' üründe maliyet yok — net kâra katılmadı (aşağıda sarı satırlara gir)</b>'):''}</div></div></div>
  <div class="card"><h2>Ürün Dağılımı <span style="text-transform:none;letter-spacing:0">· <span class="g">■</span> net kâr · <span style="color:#94a3b8">■</span> ürün maliyeti · <span class="r">■</span> komisyon+stopaj · <span class="w">■</span> kargo</span></h2>${bars}</div>
  <div class="card" style="padding:8px 16px"><table>
  <tr><th>Ürün</th><th style="text-align:right">Adet</th><th style="text-align:right">Ciro</th>
  <th style="text-align:right">Kom+Stopaj</th><th style="text-align:right">Kargo+Hizmet</th>
  <th style="text-align:right">Maliyet</th><th style="text-align:right">Net Kâr</th><th style="text-align:right">Marj</th></tr>`;
  for(const x of r.satirlar){
    if(cAra && !((x.ad||'')+x.barkod).toLocaleLowerCase('tr').includes(cAra)) continue;
    const eks=x.maliyet==null;
    h+=`<tr ${eks?'style="background:#fffbeb"':''}><td><b>${(x.ad||x.barkod).slice(0,44)}</b><div class="mut" style="font-size:10px">${x.barkod}${x.mal_birim!=null?' · birim '+tlm(x.mal_birim):''}</div></td>
    <td style="text-align:right">${x.adet}</td><td style="text-align:right;font-weight:750">${tlm(x.ciro)}</td>
    <td style="text-align:right" class="r">${tlm(x.kom_all)}${x.kom_t>0.01?'<span class="mut"> ≈</span>':''}</td>
    <td style="text-align:right" class="w">${tlm(x.kargo+(x.platform||0))}${x.kargo_t>0.01?'<span class="mut"> ≈</span>':''}</td>
    <td style="text-align:right" class="mut">${eks?'':tlm(x.maliyet)}</td>
    ${eks?`<td colspan="2" style="text-align:right"><span class="mut" style="font-size:11px">birim maliyet giriniz →</span>
      <input type="number" id="mi-${x.barkod}" placeholder="0,00" style="width:100px;padding:5px 8px;margin:0 6px"
       onkeydown="if(event.key==='Enter')malKaydet('${x.barkod}')">
      <button class="btn" style="padding:5px 12px" onclick="malKaydet('${x.barkod}')">Kaydet</button></td>`
    :`<td style="text-align:right;font-weight:850" class="${x.net>=0?'g':'r'}">${tlm(x.net)}</td>
      <td style="text-align:right;font-weight:750" class="${x.marj>=0?'g':'r'}">${pc(x.marj)}</td>`}</tr>`;}
  h+='</table></div>';
  document.getElementById('ciro').innerHTML=h;
}
async function faturaCek(){const b=document.getElementById('fbtn');b.textContent='⏳ Faturalar okunuyor...';b.disabled=true;
  const r=JSON.parse(await window.pywebview.api.fatura_maliyet());S=r.durum;ciz();
  b.textContent='📄 Fatura Maliyetleri';b.disabled=false;
  const d=r.sonuc;
  modalGoster(`<b style="color:var(--yes)">✓ Fatura taraması bitti</b>
  <div style="margin-top:10px;font-size:13px;line-height:1.9">${d.kalem} fatura kalemi · ${d.eslesen} ürün eşleşti ·
  <b>${d.yazilan} üründe maliyet güncellendi</b></div>
  <div id="log" style="margin-top:10px;max-height:220px">${(r.log||'').replace(/</g,'&lt;')}</div>`);
  window._ciroData=null;}
async function guncelle(){const b=document.getElementById('gbtn');b.textContent='⏳ Güncelleniyor...';b.disabled=true;
  const r=JSON.parse(await window.pywebview.api.guncelle());S=r.durum;ciz();
  b.textContent='⟳ Güncelle';b.disabled=false;}   // günlük ekranda değil, Trendyol/_guncelleme.log dosyasında
window.addEventListener('pywebviewready',()=>{yukle();gorunumSec('karlilik');});
</script></body></html>'''

if __name__=='__main__':
    api=Api()
    webview.create_window('Trendyol Kârlılık Paneli', html=HTML, js_api=api,
                          width=1380, height=860, min_size=(1100,700))
    webview.start()
